import calendar
import json
import locale
import math
import tempfile
from django.utils.timezone import localtime

import requests
from decimal import Decimal
from io import BytesIO
from django.template.loader import render_to_string
from weasyprint import HTML
from django.contrib.sites.shortcuts import get_current_site
import pyodbc as pyodbc
import telegram
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import AbstractUser
from django.db import connections
from django.db.models import Count, Sum, F, Q, Max
from django.utils.decorators import method_decorator
from django.contrib.auth.forms import AuthenticationForm
from django.core.paginator import Paginator
from urllib.parse import urlencode
# from django.contrib.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.edit import FormMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound, Http404, HttpResponseRedirect, JsonResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.generic.edit import FormView
from docxtpl import *
import os
from datetime import *
from number_to_string import get_string_by_number
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
import pandas as pd
from django.urls import reverse
from openpyxl.workbook import Workbook
from telegram import InlineKeyboardMarkup, InlineKeyboardButton, Update, Bot
from telegram.ext import CallbackQueryHandler, Updater, updater, Dispatcher, MessageHandler, Filters
from ekc.models import *
from ktscrm import settings
from pult.models import *
from .forms import ExcelImportForm
from .models import *
import numpy as np
from django.utils import timezone
from datetime import datetime
from django.utils.dateparse import parse_date
import openpyxl
from openpyxl.utils import get_column_letter
from .forms import *
from .models import *
from django.db.models import CharField, F, ExpressionWrapper, Value
from django.db.models import F, Value, Case, When
# from .utils import *

menu = ["О сайте", "Добавить статью", "Обратная связь", "Войти"]


def logout_view(request):
    logout(request)
    return redirect('login')


# №2 Шаблон главной страницы
@login_required
def index(request):
    context = {
        'menu': menu,
        'title': 'Главная страница'
    }
    return render(request, 'dogovornoy/index.html', context=context)


def format_date(date_str):
    try:
        # Попробуем распарсить дату из строки в формате 'гггг-мм-дд'
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        # Возвращаем дату в формате 'дд.мм.гггг'
        return parsed_date.strftime('%d.%m.%Y')
    except ValueError:
        # Если формат не соответствует, возвращаем исходную строку
        return date_str


def get_current_month_russian():
    months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    now = datetime.now()
    return months[now.month - 2]


def get_current_year():
    # Получаем текущий год
    current_year = datetime.now().year
    return current_year


@login_required
def create_dogovor(request, klient_id):
    if request.method == "GET":
        passport_info = kts.objects.get(pk=klient_id)
        vid_sign1 = vid_sign.objects.get(pk=passport_info.vid_sign_id)
        if (passport_info.urik == False) and (vid_sign1.name_sign == 'ТС'):
            doc = DocxTemplate(os.path.abspath('media/dogovor4.docx'))
            split_names_klient_ = passport_info.klient_name.split()
            if len(split_names_klient_) == 3:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.{split_names_klient_[2][0]}.'.title()
            elif len(split_names_klient_) == 2:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.'.title()
            else:
                short_names_klient = split_names_klient_[0].title()
        elif (passport_info.urik == True) and (vid_sign1.name_sign == 'ТС'):
            doc = DocxTemplate(os.path.abspath('media/dogovor5.docx'))
            short_names_klient = ""
        elif ((passport_info.mat_otv == '0') and (passport_info.urik == False)):
            doc = DocxTemplate(os.path.abspath('media/ots-fizlica-bezmat.docx'))
            split_names_klient_ = passport_info.klient_name.split()
            if len(split_names_klient_) == 3:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.{split_names_klient_[2][0]}.'.title()
            elif len(split_names_klient_) == 2:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.'.title()
            else:
                short_names_klient = split_names_klient_[0].title()
        elif ((passport_info.mat_otv == '0') and (passport_info.urik == True)):
            doc = DocxTemplate(os.path.abspath('media/dogovor3.docx'))
            short_names_klient = ""
        elif ((passport_info.mat_otv != '0') and (passport_info.urik == False)):
            doc = DocxTemplate(os.path.abspath('media/dogovor1.docx'))
            split_names_klient_ = passport_info.klient_name.split()
            if len(split_names_klient_) == 3:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.{split_names_klient_[2][0]}.'.title()
            elif len(split_names_klient_) == 2:
                short_names_klient = f'{split_names_klient_[0]} {split_names_klient_[1][0]}.'.title()
            else:
                short_names_klient = split_names_klient_[0].title()
        elif (passport_info.mat_otv != '0') and (passport_info.urik == True):
            doc = DocxTemplate(os.path.abspath('media/dogovor2.docx'))
            short_names_klient = ""
        else:
            print('test')

        additional_services_cost = passport_info.additional_services.aggregate(total_cost=Sum('price'))['total_cost']
        itog_oplata = passport_info.abon_plata + (additional_services_cost or 0)
        rekvizity_test = rekvizity.objects.get(pk=passport_info.company_name_id)
        current_date = date.today()
        current_date = current_date.strftime('%d.%m.%Y')
        now_year = datetime.now().year
        formatted_date = format_date(passport_info.data_zakluchenia)
        currency_main = ('тенге', 'тенге', 'тенге')
        currency_additional = ('тиын', 'тиына', 'тиынов')
        itog_oplata_propis = get_string_by_number(itog_oplata, currency_main, currency_additional)
        time_reag_propis = get_string_by_number(passport_info.time_reag, currency_main, currency_additional)
        oplata_itog1 = itog_oplata_propis.split('тенге 00 тиынов')
        time_reag_itog1 = time_reag_propis.split(' тенге 00 тиынов')
        time_reag_nebol_propis = get_string_by_number(passport_info.time_reag_nebol, currency_main, currency_additional)
        time_reag_nebol_itog1 = time_reag_nebol_propis.split(' тенге 00 тиынов')
        mat_otv_propis = get_string_by_number(passport_info.mat_otv, currency_main, currency_additional)
        mat_otv_itog1 = mat_otv_propis.split(' тенге 00 тиынов')
        mat_otv_itog2 = mat_otv_itog1[0].strip()

        if passport_info.email:
            email_itog = passport_info.email
        else:
            email_itog = ' '

        if passport_info.urik_adress:
            urik_adess = passport_info.urik_adress
        else:
            urik_adess = ' '

        context = {
            'snames_klient': short_names_klient,
            'udv_number': passport_info.udv_number,
            'date_udv': passport_info.date_udv,
            'dogovor_number': passport_info.dogovor_number,
            'date': current_date,
            'now_year':now_year,
            'date_zakl': formatted_date,
            'klient_name': passport_info.klient_name,
            'company_name': passport_info.company_name,
            'time_reag': passport_info.time_reag,
            'time_reag_itog1': time_reag_itog1[0],
            'time_reag_nebol': passport_info.time_reag_nebol,
            'time_reag_nebol_itog1': time_reag_nebol_itog1[0],
            'adres': passport_info.adres,
            'iin_bin': passport_info.iin_bin,
            'telephone': passport_info.telephone,
            'vid_sign_polnoe': vid_sign1.name_sign_polnoe,
            'vid_sign_sokr': vid_sign1.name_sign,
            'urik': passport_info.urik,
            'email': email_itog,
            'name_object': passport_info.name_object,
            'stoimost_rpo': passport_info.stoimost_rpo,
            'mat_otv': (passport_info.mat_otv or 0),
            'mat_otv_itog1': mat_otv_itog2,
            'itog_oplata': itog_oplata,
            'itog_oplata_propis': oplata_itog1[0],
            'chasi_po_dog': passport_info.chasi_po_dog,
            'vid_rpo': passport_info.vid_rpo,
            'polnoe_name': rekvizity_test.polnoe_name,
            'adres_company': rekvizity_test.adres_company,
            'bin': rekvizity_test.bin,
            'iban': rekvizity_test.iban,
            'bic': rekvizity_test.bic,
            'bank': rekvizity_test.bank,
            'telephone_ofiice': rekvizity_test.telephone_ofiice,
            'telephone_buh': rekvizity_test.telephone_buh,
            'vid_too': rekvizity_test.vid_too,
            'doljnost': rekvizity_test.doljnost,
            'ucheriditel_name_polnoe': rekvizity_test.ucheriditel_name_polnoe,
            'ucheriditel_name_sokr': rekvizity_test.ucheriditel_name_sokr,
            'iik': passport_info.iik,
            'bik': passport_info.bik,
            'bank_klient': passport_info.bank,
            'rezhim_raboti': passport_info.rezhim_raboti,
            'fio_direktor_sokr': passport_info.fio_direktor_sokr,
            'fio_direktor_polnoe': passport_info.fio_direktor_polnoe,
            'dolznost_klient': passport_info.dolznost,
            'ucereditel_doc': passport_info.ucereditel_doc,
            'urik_adess': urik_adess,
        }
        doc.render(context)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename=download.docx'
        doc.save(response)

    return response


@method_decorator(login_required, name='dispatch')
class AddClient(FormView):
    template_name = 'dogovornoy/add_client.html'
    form_class = AddKlientDogForm
    success_url = '/baza_dogovorov/'

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form, menu=menu, title='Новый клиент'))


@method_decorator(login_required, name='dispatch')
class AddClientPartner(CreateView):
    template_name = 'dogovornoy/add_client_partner.html'
    form_class = AddKlientDogFormPartner
    success_url = reverse_lazy('baza_partnerov')

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))



@login_required
def update_client(request, klient_id):
    kartochka = get_object_or_404(kts, pk=klient_id)
    if request.method == 'POST':
        form = AddKlientDogForm(request.POST, request.FILES, instance=kartochka)
        if form.is_valid():
            form.save()
            return redirect('kartochka_klienta', klient_id=kartochka.pk)
    else:
        form = AddKlientDogForm(instance=kartochka)

    return render(request, 'dogovornoy/update_client.html', {'form': form, 'kartochka': kartochka})


@login_required
def update_client_partner(request, partner_klient_id):
    kartochka_partner = get_object_or_404(partners_object, pk=partner_klient_id)
    if request.method == 'POST':
        form = AddKlientDogFormPartner(request.POST, request.FILES, instance=kartochka_partner)
        if form.is_valid():
            form.save()
            return redirect('baza_partnerov')
    else:
        form = AddKlientDogFormPartner(instance=kartochka_partner)

    return render(request, 'dogovornoy/update_client_partner.html', {'form': form, 'kartochka_partner': kartochka_partner})


@login_required
def delete_client(request, klient_id):
    kartochka = get_object_or_404(kts, pk=klient_id)
    if request.method == 'POST':
        kartochka.delete()
        return redirect('baza_dogovorov')
    return render(request, 'dogovornoy/delete_client.html', {'kartochka': kartochka})


@login_required
def delete_client_partners(request, partner_klient_id):
    kartochka_partner = get_object_or_404(partners_object, pk=partner_klient_id)
    if request.method == 'POST':
        kartochka_partner.delete()
        return redirect('baza_partnerov')
    return render(request, 'dogovornoy/delete_client.html', {'kartochka_partner': kartochka_partner})


@method_decorator(login_required, name='dispatch')
class DogBaza(ListView):
    model = kts
    template_name = 'dogovornoy/baza_dogovorov.html'
    context_object_name = 'klienty'

    def get(self, request, *args, **kwargs):
        # Логика фильтрации и пагинации
        queryset = kts.objects.all()
        query = self.request.GET.get('q')
        company_names = rekvizity.objects.values_list('id', 'polnoe_name')
        object_number = self.request.GET.get('object_number')
        company_name = self.request.GET.get('company_name')
        dogovor_number = self.request.GET.get('dogovor_number')
        gruppa_reagirovania = self.request.GET.get('gruppa_reagirovania')

        if query:
            queryset = queryset.filter(
                Q(object_number__icontains=query) |
                Q(dogovor_number__icontains=query) |
                Q(klient_name__icontains=query) |
                Q(adres__icontains=query) |
                Q(telephone__icontains=query) |
                Q(iin_bin__icontains=query) |
                Q(name_object__icontains=query)
            )
        if object_number:
            queryset = queryset.filter(object_number__icontains=object_number)
        if company_name:
            queryset = queryset.filter(company_name_id__exact=company_name)
        if dogovor_number:
            queryset = queryset.filter(dogovor_number__icontains=dogovor_number)
        if gruppa_reagirovania:
            queryset = queryset.filter(gruppa_reagirovania__icontains=gruppa_reagirovania)

        paginator = Paginator(queryset, per_page=25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        params = request.GET.copy()
        if 'page' in params:
            del params['page']
        pagination_url = request.path + '?' + urlencode(params)

        # Передаем форму в контекст
        task_form = TaskFormDog()

        return render(request, self.template_name, {
            'klienty': page_obj,
            'company_names': company_names,
            'pagination_url': pagination_url,
            'total_entries': queryset.count(),
            'task_form': task_form  # Передаем форму в шаблон
        })

    def post(self, request, *args, **kwargs):
        # Получаем ID клиента из POST данных формы
        client_id = request.POST.get('client_id')
        print(client_id)

        try:
            client = kts.objects.get(pk=client_id)
        except kts.DoesNotExist:
            return redirect('error_page')  # Обработайте случай, когда клиент не найден

        # Обработка формы
        task_form = TaskFormDog(request.POST)
        if task_form.is_valid():
            task = task_form.save(commit=False)
            task.created_by = request.user  # Устанавливаем текущего пользователя как создателя задачи
            task.client = client  # Устанавливаем клиента
            task.save()  # Сохраняем задачу
            return redirect('task_list')  # Перенаправляем после успешного создания задачи
        else:
            # Логика фильтрации и пагинации при ошибке
            queryset = kts.objects.all()
            paginator = Paginator(queryset, per_page=25)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            params = request.GET.copy()
            if 'page' in params:
                del params['page']
            pagination_url = request.path + '?' + urlencode(params)

            return render(request, self.template_name, {
                'klienty': page_obj,
                'task_form': task_form,  # Передаем форму с ошибками
                'pagination_url': pagination_url
            })


@method_decorator(login_required, name='dispatch')
class DogBazaPartners(ListView):
    model = partners_object
    template_name = 'dogovornoy/baza_partnerov.html'
    context_object_name = 'klienty_partners'

    def get(self, request, *args, **kwargs):
        queryset = partners_object.objects.all()
        query = self.request.GET.get('q')
        company_partners = partners_rekvizity.objects.values_list('id', 'polnoe_name')
        object_number = self.request.GET.get('object_number')
        company_name = self.request.GET.get('company_name')
        gsm_number = self.request.GET.get('gsm_number')

        if query:
            queryset = queryset.filter(
                Q(object_number__icontains=query) |
                Q(gsm_number__icontains=query) |
                Q(name_object__icontains=query) |
                Q(adres__icontains=query)
            )

        if object_number:
            queryset = queryset.filter(object_number__icontains=object_number)
        if company_name:
            queryset = queryset.filter(company_name_id__exact=company_name)
        if gsm_number:
            queryset = queryset.filter(gsm_number__icontains=gsm_number)

        paginator = Paginator(queryset, per_page=25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        params = request.GET.copy()
        if 'page' in params:
            del params['page']
        pagination_url = request.path + '?' + urlencode(params)

        return render(request, self.template_name,
                      {'klienty_partners': page_obj, 'company_partners': company_partners,
                       'pagination_url': pagination_url, 'total_entries': queryset.count()})


@method_decorator(login_required, name='dispatch')
class Rekvizity(ListView):
    model = rekvizity
    template_name = 'dogovornoy/rekvizity.html'
    context_object_name = 'rekvizity'


@method_decorator(login_required, name='dispatch')
class RekvizityPartners(ListView):
    model = partners_rekvizity
    context_object_name = 'partners_rekvizity'

@login_required
def importexcel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded file
            excel_file = request.FILES['excel_file']

            # Load the Excel data into a Pandas dataframe
            df = pd.read_excel(excel_file)
            df = df.replace({np.nan: None})
            # print(df)
            # Iterate over the rows of the dataframe and create kts objects
            for index, row in df.iterrows():
                kts_obj = kts(
                    udv_number=row['udv_number'],
                    date_udv=row['date_udv'],
                    dogovor_number=row['dogovor_number'],
                    data_zakluchenia=row['data_zakluchenia'],
                    nalichiye_dogovora=row['nalichiye_dogovora'],
                    mat_otv=row['mat_otv'],
                    act_ty=row['act_ty'],
                    time_reag=row['time_reag'],
                    time_reag_nebol=row['time_reag_nebol'],
                    yslovie_dogovora=row['yslovie_dogovora'],
                    klient_name=row['klient_name'],
                    name_object=row['name_object'],
                    adres=row['adres'],
                    iin_bin=row['iin_bin'],
                    telephone=row['telephone'],
                    urik=row['urik'],
                    chasi_po_dog=row['chasi_po_dog'],
                    dop_uslugi=row['dop_uslugi'],
                    abon_plata=row['abon_plata'],
                    object_number=row['object_number'],
                    peredatchik_number=row['peredatchik_number'],
                    stoimost_rpo=row['stoimost_rpo'],
                    date_podkluchenia=row['date_podkluchenia'],
                    date_otklulchenia=row['date_otklulchenia'],
                    gruppa_reagirovania=row['gruppa_reagirovania'],
                    email=row['email'],
                    vid_rpo=row['vid_rpo'],
                    primechanie=row['primechanie'],
                    agentskie=row['agentskie'],
                    photo=row['photo'],
                    prochee=row['prochee'],
                    company_name_id=row['company_name_id'],
                    vid_sign_id=row['vid_sign_id'],
                    date_izmenenia=row['date_izmenenia'],
                )
                kts_obj.save()

            # Redirect to the kts list page
            return HttpResponseRedirect(reverse('baza_dogovorov'))
    else:
        form = ExcelImportForm()

    return render(request, 'dogovornoy/importexel.html', {'form': form})


@login_required
def partnerts_importexel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded file
            excel_file = request.FILES['excel_file']

            # Load the Excel data into a Pandas dataframe
            df = pd.read_excel(excel_file)
            df = df.replace({np.nan: None})
            # print(df)
            # Iterate over the rows of the dataframe and create kts objects
            for index, row in df.iterrows():
                partners_obj = partners_object(
                    object_number=row['object_number'],
                    gsm_number=row['gsm_number'],
                    name_object=row['name_object'],
                    adres=row['adres'],
                    type_object=row['type_object'],
                    hours_mounth=row['hours_mounth'],
                    date_podkluchenia=row['date_podkluchenia'],
                    tariff_per_mounth=row['tariff_per_mounth'],
                    tehnical_services=row['tehnical_services'],
                    rent_gsm=row['rent_gsm'],
                    fire_alarm=row['fire_alarm'],
                    telemetria=row['telemetria'],
                    nabludenie=row['nabludenie'],
                    sms_uvedomlenie=row['sms_uvedomlenie'],
                    kolvo_day=row['kolvo_day'],
                    primechanie=row['primechanie'],
                    urik=row['urik'],
                    company_name_id=row['company_name_id'],
                    ekipazh_id=row['ekipazh_id'],
                    vid_sign_id=row['vid_sign_id'],
                    date_otkluchenia=row['date_otkluchenia'],
                )
                partners_obj.save()

            # Redirect to the kts list page
            return HttpResponseRedirect(reverse('baza_partnerov'))
    else:
        form = PartnersImportForm()

    return render(request, 'dogovornoy/partnerts_importexel.html', {'form': form})


@login_required
def importrekvizity(request):
    if request.method == 'POST':
        form = RekvizityImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded file
            excel_file = request.FILES['excel_file']

            # Load the Excel data into a Pandas dataframe
            df = pd.read_excel(excel_file)
            df = df.replace({np.nan: None})
            # print(df)
            # Iterate over the rows of the dataframe and create kts objects
            for index, row in df.iterrows():
                rekvizity_obj = rekvizity(
                    id=row['id'],
                    polnoe_name=row['polnoe_name'],
                    adres_company=row['adres_company'],
                    bin=row['bin'],
                    iban=row['iban'],
                    bic=row['bic'],
                    bank=row['bank'],
                    telephone_ofiice=row['telephone_ofiice'],
                    telephone_buh=row['telephone_buh'],
                    vid_too=row['vid_too'],
                    doljnost=row['doljnost'],
                    ucheriditel_name_polnoe=row['ucheriditel_name_polnoe'],
                    ucheriditel_name_sokr=row['ucheriditel_name_sokr'],
                )
                rekvizity_obj.save()

            # Redirect to the kts list page
            return HttpResponseRedirect(reverse('rekvizity'))
    else:
        form = RekvizityImportForm()

    return render(request, 'dogovornoy/importrekvizity.html', {'form': form})


@login_required
def importvidsign(request):
    if request.method == 'POST':
        form = VidSignImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded file
            excel_file = request.FILES['excel_file']

            # Load the Excel data into a Pandas dataframe
            df = pd.read_excel(excel_file)
            df = df.replace({np.nan: None})
            # print(df)
            # Iterate over the rows of the dataframe and create kts objects
            for index, row in df.iterrows():
                vidsign_obj = vid_sign(
                    id=row['id'],
                    name_sign=row['name_sign'],
                    name_sign_polnoe=row['name_sign_polnoe'],
                )
                vidsign_obj.save()

            # Redirect to the kts list page
            return HttpResponseRedirect(reverse('baza_dogovorov'))
    else:
        form = VidSignImportForm()

    return render(request, 'dogovornoy/importvidsign.html', {'form': form})


@login_required
def importekipazh(request):
    if request.method == 'POST':
        form = VidSignImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded file
            excel_file = request.FILES['excel_file']

            # Load the Excel data into a Pandas dataframe
            df = pd.read_excel(excel_file)
            df = df.replace({np.nan: None})
            # print(df)
            # Iterate over the rows of the dataframe and create kts objects
            for index, row in df.iterrows():
                ekipazh_obj = ekipazh(
                    id=row['id'],
                    ekipazh_name=row['ekipazh_name'],
                )
                ekipazh_obj.save()

            # Redirect to the kts list page
            return HttpResponseRedirect(reverse('baza_dogovorov'))
    else:
        form = EkipazhImportForm()

    return render(request, 'dogovornoy/importekipazh.html', {'form': form})


class CopyClientView(View):
    def get(self, request, pk):
        # Находим оригинального клиента
        original_client = get_object_or_404(kts, pk=pk)

        # Проверяем, есть ли у оригинального клиента обязательные поля
        if not original_client.company_name:
            # Вы можете либо установить значение по умолчанию, либо перенаправить с сообщением об ошибке
            return redirect('error_page')  # Убедитесь, что вы обрабатываете этот случай

        # Копируем все данные, кроме ID и обязательные поля заполняем корректно
        new_client = kts.objects.create(
            udv_number=original_client.udv_number,
            date_udv = original_client.date_udv,
            company_name = original_client.company_name,
            dogovor_number = original_client.dogovor_number,
            data_zakluchenia = original_client.data_zakluchenia,
            nalichiye_dogovora = original_client.nalichiye_dogovora,
            mat_otv = original_client.mat_otv,
            act_ty = original_client.act_ty,
            time_reag = original_client.time_reag,
            time_reag_nebol = original_client.time_reag_nebol,
            yslovie_dogovora = original_client.yslovie_dogovora,
            klient_name = original_client.klient_name,
            name_object = original_client.name_object,
            adres = original_client.adres,
            iin_bin = original_client.iin_bin,
            telephone = original_client.telephone,
            vid_sign = original_client.vid_sign,
            urik = original_client.urik,
            chasi_po_dog = original_client.chasi_po_dog,
            dop_uslugi = original_client.dop_uslugi,
            abon_plata = original_client.abon_plata,
            object_number = original_client.object_number,
            peredatchik_number = original_client.peredatchik_number,
            stoimost_rpo = original_client.stoimost_rpo,
            date_podkluchenia = original_client.date_podkluchenia,
            date_otklulchenia = original_client.date_otklulchenia,
            date_izmenenia = original_client.date_izmenenia,
            gruppa_reagirovania = original_client.gruppa_reagirovania,
            email = original_client.email,
            vid_rpo = original_client.vid_rpo,
            primechanie = original_client.primechanie,
            agentskie = original_client.agentskie,
            photo = original_client.photo,
            prochee = original_client.prochee,
            exclude_from_report = original_client.exclude_from_report,
        )

        # Перенаправляем на страницу базы договоров после успешного копирования
        return redirect('update_client', klient_id=new_client.pk)



class CopyClientViewPartner(View):
    def get(self, request, pk):
        # Находим оригинального клиента
        original_client = get_object_or_404(partners_object, pk=pk)

        # Проверяем, есть ли у оригинального клиента обязательные поля, которые нужно скопировать корректно
        if not original_client.company_name:
            # Устанавливаем значение по умолчанию или перенаправляем с сообщением об ошибке
            return redirect('error_page')

        # Копируем все данные, кроме ID, и сохраняем нового клиента
        new_client = partners_object.objects.create(
            object_number=original_client.object_number,
            gsm_number=original_client.gsm_number,
            name_object=original_client.name_object,
            adres=original_client.adres,
            type_object=original_client.type_object,
            vid_sign=original_client.vid_sign,
            hours_mounth=original_client.hours_mounth,
            date_podkluchenia=original_client.date_podkluchenia,
            tariff_per_mounth=original_client.tariff_per_mounth,
            tehnical_services=original_client.tehnical_services,
            rent_gsm=original_client.rent_gsm,
            fire_alarm=original_client.fire_alarm,
            telemetria=original_client.telemetria,
            nabludenie=original_client.nabludenie,
            sms_uvedomlenie=original_client.sms_uvedomlenie,
            sms_number=original_client.sms_number,
            kolvo_day=original_client.kolvo_day,
            primechanie=original_client.primechanie,
            ekipazh=original_client.ekipazh,
            urik=original_client.urik,
            company_name=original_client.company_name,
            date_otkluchenia=original_client.date_otkluchenia,
            prochee=original_client.prochee
        )

        # Перенаправляем на страницу нового клиента
        return redirect('update_client_partner', partner_klient_id=new_client.pk)



def calculate_active_days(partner_object, start_of_month=None, end_of_month=None):
    now = timezone.now()
    if not start_of_month or not end_of_month:
        previous_month = now.month or 12
        year = now.year if now.month > 1 else now.year - 1
        start_of_month = datetime(year, previous_month, 1, tzinfo=timezone.utc).date()
        end_of_month = datetime(year, previous_month, calendar.monthrange(year, previous_month)[1], tzinfo=timezone.utc).date()
        print(start_of_month)
        print(end_of_month)

    if not partner_object.date_podkluchenia:
        return 0

    podkluchenia = partner_object.date_podkluchenia

    # Проверяем наличие поля даты отключения и извлекаем его
    if hasattr(partner_object, 'date_otkluchenia') and partner_object.date_otkluchenia:
        otkluchenia = partner_object.date_otkluchenia
    elif hasattr(partner_object, 'date_otklulchenia') and partner_object.date_otklulchenia:
        otkluchenia = partner_object.date_otklulchenia
    else:
        otkluchenia = None


    if not otkluchenia:
        if podkluchenia <= start_of_month:
            return (end_of_month - start_of_month).days + 1
        return (end_of_month - podkluchenia).days + 1

    if otkluchenia < start_of_month:
        return 0

    if podkluchenia > end_of_month:
        return 0

    effective_start = max(podkluchenia, start_of_month)
    effective_end = min(otkluchenia, end_of_month)

    return (effective_end - effective_start).days + 1



def calculate_service_active_days(service_instance, start_of_month=None, end_of_month=None):
    now = timezone.now()

    # Установить начало и конец месяца, если они не указаны
    if not start_of_month or not end_of_month:
        previous_month = now.month - 1 or 12
        year = now.year if now.month > 1 else now.year - 1
        start_of_month = datetime(year, previous_month, 1, tzinfo=timezone.utc).date()
        end_of_month = datetime(year, previous_month, calendar.monthrange(year, previous_month)[1], tzinfo=timezone.utc).date()

    # Проверить, подключена ли услуга
    if not service_instance.date_added:
        return 0

    date_added = service_instance.date_added
    date_unsubscribe = service_instance.date_unsubscribe

    # Если услуга никогда не была отключена
    if not date_unsubscribe:
        if date_added <= start_of_month:
            return (end_of_month - start_of_month).days + 1
        return (end_of_month - date_added).days + 1

    # Если услуга была отключена до начала месяца
    if date_unsubscribe < start_of_month:
        return 0

    # Если услуга была подключена после конца месяца
    if date_added > end_of_month:
        return 0

    # Рассчитать эффективные даты подключения
    effective_start = max(date_added, start_of_month)
    effective_end = min(date_unsubscribe, end_of_month)

    return (effective_end - effective_start).days + 1




@method_decorator(login_required, name='dispatch')
class KartochkaKlienta(DetailView):
    model = kts
    template_name = 'dogovornoy/kartochka_klienta.html'
    pk_url_kwarg = 'klient_id'
    context_object_name = 'kartochka'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        kts_instance = self.get_object()

        now = timezone.now()
        num_days_mounth = calendar.monthrange(now.year, now.month)[1]
        start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc).date()
        end_of_month = datetime(now.year, now.month, num_days_mounth, tzinfo=timezone.utc).date()
        # Используем функцию calculate_active_days для расчета num_days
        num_days = calculate_active_days(kts_instance, start_of_month, end_of_month)
        print(num_days)

        total_additional_services_cost = 0
        itog_abon_plata = 0
        services_with_cost = []

        for service in kts_instance.additional_services.all():
            active_days = calculate_service_active_days(service, start_of_month, end_of_month)
            service_cost = (service.price / num_days_mounth) * active_days if active_days > 0 else 0

            total_additional_services_cost += service_cost

            services_with_cost.append ({
                'id': service.pk,
                'service_name': service.service_name,
                'date_added': service.date_added,
                'date_unsubscribe': service.date_unsubscribe,
                'active_days': active_days,
                'price_per_month': service.price,
                'calculated_cost': service_cost
            })

        # Преобразование abon_plata в Decimal
        itog_abon_plata = Decimal(kts_instance.abon_plata) / Decimal(num_days_mounth) * Decimal(num_days)

        # Итоговая оплата
        itog_oplata = itog_abon_plata + total_additional_services_cost
        itog_oplata = round(itog_oplata, 2)

        context['itog_oplata'] = itog_oplata
        context['num_days'] = num_days
        context['services_with_cost'] = services_with_cost
        context['total_additional_services_cost'] = round(total_additional_services_cost, 2)
        context['form'] = AdditionalServiceForm()
        return context

    def post(self, request, *args, **kwargs):
        form = AdditionalServiceForm(request.POST)
        if form.is_valid():
            kts_instance = self.get_object()  # Get the Kts instance associated with the view
            form.instance.kts = kts_instance  # Associate the additional service with the Kts instance
            form.save()  # Save the additional service
            return redirect('kartochka_klienta', klient_id=kts_instance.pk)
        else:
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)


def format_date_russian(date):
    months = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
        7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
    }
    return f"{date.day} {months[date.month]} {date.year}"


def format_date_russian_invoice(date):
    months = {
        1: "январе", 2: "феврале", 3: "марте", 4: "апреле", 5: "мае", 6: "июне",
        7: "июле", 8: "августе", 9: "сентябре", 10: "октябре", 11: "ноябре", 12: "декабре"
    }
    return f"{months[date.month]} {date.year}"


def save_pdf_to_invoices(html_string, invoice_number):
    # Путь к папке invoices в MEDIA_ROOT
    invoices_dir = os.path.join(settings.MEDIA_ROOT, 'invoices')
    os.makedirs(invoices_dir, exist_ok=True)  # Создаем папку, если она не существует

    # Путь для сохранения файла
    file_name = f'invoice_{invoice_number}.pdf'
    file_path = os.path.join(invoices_dir, file_name)

    # Генерация PDF и сохранение
    with open(file_path, 'wb') as pdf_file:
        HTML(string=html_string).write_pdf(pdf_file)

    return file_path, file_name


def send_whatsapp_pdf(phone_number, pdf_path, access_token, message, channel_id):
    # Отправляем документ через WhatsApp API
    url = f"https://api.wazzup24.com/v3/message"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }


    # Создаем публичный URL для временного файла
    # (вместо этого вы можете загрузить файл на ваш сервер или S3 и получить публичный URL)
    public_url = "https://disk.yandex.ru/i/c7T5lpf9sJkhMw"  # Публичный URL временного файла

    # Запрос на отправку сообщения через WhatsApp
    data = {
        "channelId": channel_id,
        "chatType": "whatsapp",
        "chatId": phone_number,  # Номер клиента в международном формате
        "contentUri": pdf_path,
    }

    response = requests.post(url, headers=headers, json=data)

    # Проверяем статус запроса
    if response.status_code == 201:
        print("Сообщение успешно отправлено!")
        return True
    else:
        print(f"Ошибка: {response.status_code} - {response.json()}")
        return False


def send_whatsapp_message(phone_number, pdf_path, access_token, message, channel_id):
    # Отправляем документ через WhatsApp API
    url = f"https://api.wazzup24.com/v3/message"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }


    # Создаем публичный URL для временного файла
    # (вместо этого вы можете загрузить файл на ваш сервер или S3 и получить публичный URL)
    public_url = "https://disk.yandex.ru/i/c7T5lpf9sJkhMw"  # Публичный URL временного файла

    # Запрос на отправку сообщения через WhatsApp
    data = {
        "channelId": channel_id,
        "chatType": "whatsapp",
        "chatId": phone_number,  # Номер клиента в международном формате
        "text": message,
    }

    response = requests.post(url, headers=headers, json=data)

    # Проверяем статус запроса
    if response.status_code == 201:
        print("Сообщение успешно отправлено!")
        return True
    else:
        print(f"Ошибка: {response.status_code} - {response.json()}")
        return False


def generate_invoice(request, pk):
    # Получаем клиента
    kts_instance = kts.objects.get(pk=pk)

    # Получаем данные по услугам
    services_with_cost = []
    total_services_cost = 0
    now = timezone.now()
    formatted_date = format_date_russian(now)
    formatted_date_month = format_date_russian_invoice(now)
    num_days_mounth = calendar.monthrange(now.year, now.month)[1]
    start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month, num_days_mounth, tzinfo=timezone.utc).date()
    kbe = 17 if "ТОО" in kts_instance.company_name.polnoe_name else 19
    company = kts_instance.company_name

    for service in kts_instance.additional_services.all():
        active_days = calculate_service_active_days(service, start_of_month, end_of_month)
        service_cost = (service.price / num_days_mounth) * active_days if active_days > 0 else 0
        total_services_cost += service_cost
        services_with_cost.append({
            'service_name': service.service_name,
            'active_days': active_days,
            'price_per_month': service.price,
            'calculated_cost': service_cost
        })

    # Расчет итоговой суммы
    num_days = calculate_active_days(kts_instance, start_of_month, end_of_month)
    abon_plata = Decimal(kts_instance.abon_plata) / Decimal(num_days_mounth) * Decimal(num_days)
    total_cost = abon_plata + total_services_cost
    nds = (total_cost / 100) * 13
    currency_main = ('тенге', 'тенге', 'тенге')
    currency_additional = ('тиын', 'тиына', 'тиынов')
    itog_oplata_propis = get_string_by_number(total_cost, currency_main, currency_additional)

    # Найти максимальный номер счета
    last_number = Invoice.objects.aggregate(Max('number'))['number__max'] or 0

    # Увеличить номер на 1
    new_number = last_number + 1

    # Создать новый счет
    invoice = Invoice.objects.create(
        number=new_number,
        client=kts_instance.company_name.polnoe_name,
        total_amount=total_cost
    )

    current_site = get_current_site(request)
    company_seal_url = request.build_absolute_uri(company.img_pechat.url) if company.img_pechat else None

    # Подготовка данных для шаблона
    context = {
        'client_name': kts_instance.company_name.polnoe_name,
        'adres_company': kts_instance.company_name.adres_company,
        'telephone_ofiice': kts_instance.company_name.telephone_ofiice,
        'kbe': kbe,
        'bin': kts_instance.company_name.bin,
        'iban': kts_instance.company_name.iban,
        'bank': kts_instance.company_name.bank,
        'bic': kts_instance.company_name.bic,
        'invoice_date': formatted_date,
        'last_number': last_number,
        'invoice_date_month': formatted_date_month,
        'klient_name': kts_instance.klient_name,
        'iin_bin': kts_instance.iin_bin,
        'adres': kts_instance.adres,
        'telephone': kts_instance.telephone,
        'itog_oplata_propis': itog_oplata_propis,
        'services': services_with_cost,
        'abon_plata': round(abon_plata, 2),
        'total_cost': round(total_cost, 2),
        'nds': round(nds, 2),
        'company_seal': company_seal_url,
    }

    # Генерация HTML из шаблона
    html_string = render_to_string('dogovornoy/invoice.html', context)

    # Сохранение PDF в папку invoices
    file_path, file_name = save_pdf_to_invoices(html_string, new_number)
    file_url = f"https://kateryushin.pro/{settings.MEDIA_URL}invoices/{file_name}"  # Публичный URL

    # Отправка через WhatsApp
    access_token = "f895ca7a98494aa6b1dd7a4cab83f026"
    channel_id = "da3aa85a-4133-44a5-8e4c-1c259e0fb885"
    phone_number = kts_instance.telephone
    message = f"Здравствуйте, {kts_instance.company_name.polnoe_name}! Ваш счет на оплату готов. Общая сумма: {kts_instance.abon_plata} тенге."
    send_whatsapp_message(phone_number, file_url, access_token, message, channel_id)
    send_whatsapp_pdf(phone_number, file_url, access_token, message, channel_id)

    # Возврат HTTP-ответа
    with open(file_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{file_name}"'

    return response



@method_decorator(login_required, name='dispatch')
class KartochkaPartner(DetailView):
    model = partners_object
    template_name = 'dogovornoy/kartochka_partner.html'
    pk_url_kwarg = 'partner_klient_id'
    context_object_name = 'kartochka_partner'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        now = timezone.now()
        partner_object = self.get_object()
        num_days_mounth = calendar.monthrange(now.year, now.month)[1]
        start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc).date()
        end_of_month = datetime(now.year, now.month, num_days_mounth, tzinfo=timezone.utc).date()

        # Используем функцию calculate_active_days для расчета num_days
        num_days = calculate_active_days(partner_object, start_of_month, end_of_month)
        print(num_days)

        sms_uvedomlenie = 0

        # Calculate itog_tehnical_services
        if partner_object.tehnical_services:
            if partner_object.urik:
                itog_tehnical_services = int(
                    (partner_object.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days)
            else:
                itog_tehnical_services = int(
                    (partner_object.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days)
        else:
            itog_tehnical_services = 0


        if partner_object.rent_gsm:
            if partner_object.urik:
                itog_rent_gsm = int((partner_object.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((partner_object.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0


        if partner_object.fire_alarm:
            if partner_object.urik:
                itog_fire_alarm = int((partner_object.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((partner_object.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0


        if partner_object.telemetria:
            itog_telemetria = int((partner_object.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if partner_object.nabludenie:
            if partner_object.urik:
                itog_nabludenie = int((partner_object.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((partner_object.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if partner_object.sms_uvedomlenie:
            sms_uvedomlenie = partner_object.company_name.sms
            if partner_object.sms_number:
                itog_sms_uvedomlenie = int(
                    ((partner_object.company_name.sms * partner_object.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((partner_object.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0


        # Calculate reagirovanie
        if partner_object.urik:
            reagirovanie = (partner_object.hours_mounth * partner_object.tariff_per_mounth) / num_days_mounth * num_days
        else:
            reagirovanie = (partner_object.tariff_per_mounth / num_days_mounth) * num_days

        reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
            reagirovanie)


        if partner_object.primechanie == '50% на 50%':
            summ_mounth = int((itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2)
        else:
            summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm


        context['itog_tehnical_services'] = itog_tehnical_services
        context['reagirovanie'] = reagirovanie
        context['num_days'] = num_days
        context['itog_rent_gsm'] = itog_rent_gsm
        context['itog_fire_alarm'] = itog_fire_alarm
        context['itog_telemetria'] = itog_telemetria
        context['itog_nabludenie'] = itog_nabludenie
        context['itog_sms_uvedomlenie'] = itog_sms_uvedomlenie
        context['sms_uvedomlenie'] = sms_uvedomlenie
        context['summ_mounth'] = summ_mounth

        return context

@login_required
def delete_additional_service(request, service_id):
    additional_service = get_object_or_404(AdditionalService, pk=service_id)

    if request.method == 'POST':
        additional_service.delete()
        return redirect('kartochka_klienta', klient_id=additional_service.kts_id)

    return render(request, 'dogovornoy/delete_additional_service.html', {'additional_service': additional_service})


@login_required
def edit_additional_service(request, service_id):
    additional_service = get_object_or_404(AdditionalService, pk=service_id)

    if request.method == 'POST':
        form = AdditionalServiceForm(request.POST, instance=additional_service)
        if form.is_valid():
            form.save()
            return redirect('kartochka_klienta', klient_id=additional_service.kts_id)
    else:
        form = AdditionalServiceForm(instance=additional_service)

    return render(request, 'dogovornoy/edit_additional_service.html',
                  {'form': form, 'additional_service': additional_service})


# Страница 404
def pageNotFound(request, exception):
    return HttpResponseNotFound('<h1>Сраница не найдена</h1>')


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            login(request, form.get_user())
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'accounts/login.html', {'form': form})



@login_required
def reports(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    companies = rekvizity.objects.all()

    reports = []

    for company in companies:
        kts_company_name = kts.objects.filter(company_name_id=company.id).distinct()

        kts_otkl = kts.objects.filter(company_name_id=company.id, date_otklulchenia__gte=start_of_month,
              date_otklulchenia__lte=end_of_month, exclude_from_report=False).distinct()

        kts_podkl = kts.objects.filter(company_name_id=company.id, date_podkluchenia__gte=start_of_month,
              date_podkluchenia__lte=end_of_month, exclude_from_report=False).distinct()

        kts_izmenenie = kts.objects.filter(
            Q(company_name_id=company.id, date_izmenenia__gte=start_of_month,
              date_izmenenia__lte=end_of_month, exclude_from_report=False) |
            Q(company_name_id=company.id, additional_services__date_added__gte=start_of_month,
              additional_services__date_added__lte=end_of_month, exclude_from_report=False) |
            Q(company_name_id=company.id, additional_services__date_unsubscribe__gte=start_of_month,
              additional_services__date_unsubscribe__lte=end_of_month, exclude_from_report=False)
        ).distinct()

        kts_abon_summa_otkl = kts_otkl.aggregate(Sum('abon_plata'))
        kts_count_otkl = kts_otkl.aggregate(Count('id'))
        kts_fiz_otkl = kts_otkl.filter(urik=False).aggregate(Count('id'))

        kts_abon_summa_podkl = kts_podkl.aggregate(Sum('abon_plata'))
        kts_count_podkl = kts_podkl.exclude(additional_services__date_added__lte=end_of_month).aggregate(Count('id'))
        kts_fiz_podkl = kts_podkl.filter(urik=False).aggregate(Count('id'))

        kts_abon_summa_izmenenia = kts_izmenenie.aggregate(Sum('abon_plata'))
        kts_count_izmenenia = kts_izmenenie.aggregate(Count('id'))
        kts_fiz_izmenenia = kts_izmenenie.filter(urik=False).aggregate(Count('id'))

        for kts_instance in kts_podkl:
            additional_services_cost = kts_instance.additional_services.aggregate(total_cost=Sum('price'))['total_cost']
            additional_services_prim = kts_instance.additional_services.all()
            for service in additional_services_prim:
                kts_instance.primechanie = f"{kts_instance.primechanie}, '{service.service_name}' с '{service.date_added}' а/п была = {kts_instance.abon_plata or 0} "

            if additional_services_cost:
                if kts_instance.date_podkluchenia.month != start_of_month.month:
                    kts_instance.abon_plata = additional_services_cost
                    kts_abon_summa_podkl['abon_plata__sum'] = (kts_abon_summa_podkl['abon_plata__sum'] or 0) + additional_services_cost
                else:
                    kts_instance.abon_plata += additional_services_cost
                    kts_abon_summa_podkl['abon_plata__sum'] = (kts_abon_summa_podkl['abon_plata__sum'] or 0) + additional_services_cost

        for kts_instance in kts_otkl:
            additional_services_cost = kts_instance.additional_services.aggregate(total_cost=Sum('price'))['total_cost']
            additional_services_prim = kts_instance.additional_services.all()
            for service in additional_services_prim:
                kts_instance.primechanie = f"{kts_instance.primechanie}, '{service.service_name}' с '{service.date_unsubscribe}' а/п была = {kts_instance.abon_plata or 0} "

            if additional_services_cost:
                if kts_instance.date_otklulchenia.month != start_of_month.month:
                    kts_instance.abon_plata = additional_services_cost
                else:
                    kts_instance.abon_plata += additional_services_cost
                    kts_abon_summa_otkl['abon_plata__sum'] = (kts_abon_summa_otkl['abon_plata__sum'] or 0) + additional_services_cost


        for kts_instance in kts_izmenenie:
            additional_services_cost = kts_instance.additional_services.aggregate(total_cost=Sum('price'))['total_cost']
            additional_services_prim = kts_instance.additional_services.all()
            for service in additional_services_prim:
                kts_instance.primechanie = f"{kts_instance.primechanie}, '{service.service_name}' с '{service.date_added}' а/п была = {kts_instance.abon_plata or 0} "

            if additional_services_cost:
                if kts_instance.date_podkluchenia.month != start_of_month.month:
                    kts_instance.abon_plata = additional_services_cost
                    kts_abon_summa_izmenenia['abon_plata__sum'] = (kts_abon_summa_izmenenia['abon_plata__sum'] or 0) + additional_services_cost
                else:
                    kts_instance.abon_plata += additional_services_cost
                    kts_abon_summa_izmenenia['abon_plata__sum'] = (kts_abon_summa_izmenenia['abon_plata__sum'] or 0) + additional_services_cost


        reports.append({
            'kts_company_name': kts_company_name,
            'company': company,
            'kts_otkl': kts_otkl,
            'kts_count_otkl': kts_count_otkl,
            'kts_fiz_otkl': kts_fiz_otkl,
            'kts_podkl': kts_podkl,
            'kts_abon_summa_podkl': kts_abon_summa_podkl,
            'kts_abon_summa_otkl': kts_abon_summa_otkl,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'start_of_month': start_of_month,
            'end_of_month': end_of_month,
            'kts_abon_summa_izmenenia':kts_abon_summa_izmenenia,
            'kts_count_izmenenia':kts_count_izmenenia,
            'kts_fiz_izmenenia':kts_fiz_izmenenia,
            'kts_izmenenie':kts_izmenenie,
        })

    context = {'reports': reports, 'start_of_month': start_of_month, 'end_of_month': end_of_month}
    return render(request, 'dogovornoy/reports.html', context)



@login_required
def export_reports_to_excel(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    companies = rekvizity.objects.all()

    reports = []

    for company in companies:
        kts_company_name = kts.objects.filter(company_name_id=company.id).distinct()

        kts_otkl = kts.objects.filter(
            Q(company_name_id=company.id, date_otklulchenia__gte=start_of_month,
              date_otklulchenia__lte=end_of_month, exclude_from_report=False) |
            Q(company_name_id=company.id, additional_services__date_unsubscribe__gte=start_of_month,
              additional_services__date_unsubscribe__lte=end_of_month, exclude_from_report=False)
        ).distinct()

        kts_podkl = kts.objects.filter(
            Q(company_name_id=company.id, date_podkluchenia__gte=start_of_month,
              date_podkluchenia__lte=end_of_month, exclude_from_report=False) |
            Q(company_name_id=company.id, additional_services__date_added__gte=start_of_month,
              additional_services__date_added__lte=end_of_month, exclude_from_report=False)
        ).distinct()

        kts_izmenenie = kts.objects.filter(
            Q(company_name_id=company.id, date_izmenenia__gte=start_of_month,
              date_izmenenia__lte=end_of_month, exclude_from_report=False) |
            Q(company_name_id=company.id, additional_services__date_added__gte=start_of_month,
              additional_services__date_added__lte=end_of_month, exclude_from_report=False)
        ).distinct()

        kts_abon_summa_otkl = kts_otkl.aggregate(Sum('abon_plata'))
        kts_count_otkl = kts_otkl.aggregate(Count('id'))
        kts_fiz_otkl = kts_otkl.filter(urik=False).aggregate(Count('id'))

        kts_abon_summa_podkl = kts_podkl.aggregate(Sum('abon_plata'))
        kts_count_podkl = kts_podkl.aggregate(Count('id'))
        kts_fiz_podkl = kts_podkl.filter(urik=False).aggregate(Count('id'))

        kts_abon_summa_izmenenia = kts_izmenenie.aggregate(Sum('abon_plata'))
        kts_count_izmenenia = kts_izmenenie.aggregate(Count('id'))
        kts_fiz_izmenenia = kts_izmenenie.filter(urik=False).aggregate(Count('id'))

        reports.append({
            'kts_company_name': kts_company_name,
            'company': company,
            'kts_otkl': kts_otkl,
            'kts_count_otkl': kts_count_otkl,
            'kts_fiz_otkl': kts_fiz_otkl,
            'kts_podkl': kts_podkl,
            'kts_abon_summa_podkl': kts_abon_summa_podkl,
            'kts_abon_summa_otkl': kts_abon_summa_otkl,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'start_of_month': start_of_month,
            'end_of_month': end_of_month,
            'kts_abon_summa_izmenenia': kts_abon_summa_izmenenia,
            'kts_count_izmenenia': kts_count_izmenenia,
            'kts_fiz_izmenenia': kts_fiz_izmenenia,
            'kts_izmenenie': kts_izmenenie,
        })

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по Экипажам"

    # Define column headers
    columns = [
        "Компания", "Тип отчета", "№ дог", "Дата", "Клиент", "Наименование объекта",
        "Адрес", "ИИН/БИН", "Вид сигнализации", "Часы по договору",
        "Алсеко", "сум.план", "№ объекта", "№ ЦМТ/GSM", "Стоимость РПО",
        "Дата подключения/отключения/изменения", "Группа реагирования", "Примечание"
    ]
    ws.append(columns)

    # Populate the worksheet with data
    for con in reports:
        for report_type, kts_list in [('Подключенные', con['kts_podkl']),
                                      ('Отключенные', con['kts_otkl']),
                                      ('Изменения', con['kts_izmenenie'])]:
            for kts_instance in kts_list:
                additional_services = ", ".join(
                    f"{service.service_name} ({service.date_added if report_type == 'Подключенные' else service.date_unsubscribe})"
                    for service in kts_instance.additional_services.all()
                )
                row = [
                    con['company'].polnoe_name,
                    report_type,
                    kts_instance.dogovor_number,
                    kts_instance.data_zakluchenia,
                    kts_instance.klient_name,
                    kts_instance.name_object,
                    kts_instance.adres,
                    kts_instance.iin_bin,
                    kts_instance.vid_sign.name_sign,  # Convert to string
                    kts_instance.chasi_po_dog,
                    additional_services,
                    kts_instance.abon_plata,
                    kts_instance.object_number,
                    kts_instance.peredatchik_number,
                    kts_instance.stoimost_rpo,
                    kts_instance.date_podkluchenia if report_type == 'Подключенные' else
                    kts_instance.date_otklulchenia if report_type == 'Отключенные' else
                    kts_instance.date_izmenenia,
                    kts_instance.gruppa_reagirovania,
                    kts_instance.primechanie
                ]
                ws.append(row)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HttpResponse object with the appropriate Excel header
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reports.xlsx'

    return response





# Страница отчеты договорной
@login_required
def reports_agentskie(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc)
    end_of_month = timezone.datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    companies = rekvizity.objects.all()
    urik_companies = rekvizity.objects.filter(kts__urik=True, kts__date_otklulchenia=None)
    non_urik_companies_quantity = rekvizity.objects.filter(kts__urik=False, kts__date_otklulchenia=None)

    reports = []

    for company in companies:
        # 1 otlk
        kts_otkl = kts.objects.filter(company_name_id=company.id, date_otklulchenia__gte=start_of_month,
                                      date_otklulchenia__lte=end_of_month)
        kts_abon_summa = kts.objects.filter(company_name_id=company.id, date_otklulchenia__gte=start_of_month,
                                            date_otklulchenia__lte=end_of_month).aggregate(Sum('abon_plata'))
        kts_count = kts.objects.filter(company_name_id=company.id, date_otklulchenia__gte=start_of_month,
                                       date_otklulchenia__lte=end_of_month).aggregate(Count('id'))
        kts_fiz = kts.objects.filter(company_name_id=company.id, urik=False, date_otklulchenia__gte=start_of_month,
                                     date_otklulchenia__lte=end_of_month).aggregate(Count('id'))
        # 2 podlk
        kts_podkl = kts.objects.filter(company_name_id=company.id, date_podkluchenia__gte=start_of_month,
                                       date_podkluchenia__lte=end_of_month)
        kts_abon_summa_podkl = kts.objects.filter(company_name_id=company.id, date_podkluchenia__gte=start_of_month,
                                                  date_podkluchenia__lte=end_of_month).aggregate(Sum('abon_plata'))
        kts_count_podkl = kts.objects.filter(company_name_id=company.id, date_podkluchenia__gte=start_of_month,
                                             date_podkluchenia__lte=end_of_month).aggregate(Count('id'))
        kts_fiz_podkl = kts.objects.filter(company_name_id=company.id, urik=False,
                                           date_podkluchenia__gte=start_of_month,
                                           date_podkluchenia__lte=end_of_month).aggregate(Count('id'))
        reports.append({
            'companies': companies,
            'urik_companies': urik_companies,
            'non_urik_companies_quantity': non_urik_companies_quantity,
            'kts_otkl': kts_otkl,
            'kts_abon_summa': kts_abon_summa,
            'kts_count': kts_count,
            'kts_fiz': kts_fiz,
            'kts_podkl': kts_podkl,
            'kts_abon_summa_podkl': kts_abon_summa_podkl,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'start_of_month': start_of_month,
            'end_of_month': end_of_month,
        })
        context = {'reports': reports, 'start_of_month': start_of_month, 'end_of_month': end_of_month}
    return render(request, 'dogovornoy/reports_agentskie.html', context)

@login_required
def reports_partners(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth


    partners_object_podkl = partners_object.objects.filter(
        company_name_id=1
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    )
    partners_kolvo_object = partners_object.objects.filter(
        company_name_id=1
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(
        company_name_id=1, urik=True
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(
        company_name_id=1, urik=False
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))

    print(end_of_month)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth


        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days)
            else:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = int(((kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)
        else:
            reagirovanie = int(((kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners.html', context)


@login_required
def reports_partners_download_urik(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=1, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1


    partners_object_podkl = partners_object.objects.filter(
        company_name_id=1, urik=True
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    )

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth


        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days)
            else:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = int(
                ((kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)
        else:
            reagirovanie = int(((kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services + summ_fire_alarm
        summ_all_company = summ_kts + summ_senim

        reports.append({
            'kts_instance': kts_instance,
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
            'summ_all_company': summ_all_company,
        })

    template_path = os.path.join(settings.MEDIA_ROOT, 'reports_partner_sgsplus_urik.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'C{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 11
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['itog_telemetria']
        ws[f'K{row_num}'] = report['itog_rent_gsm']
        ws[f'L{row_num}'] = report['itog_nabludenie']
        ws[f'M{row_num}'] = report['reagirovanie']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+3}'] = 'Итого'
    ws[f'J{row_num+3}'] = report['summ_telemetria']
    ws[f'K{row_num+3}'] = report['summ_rent_gsm']
    ws[f'L{row_num+3}'] = report['summ_nabludenie']
    ws[f'M{row_num+3}'] = report['summ_reagirovanie']
    ws[f'N{row_num+3}'] = report['summ_tehnical_services']
    ws[f'O{row_num+3}'] = report['summ_sms_uvedomlenie']
    ws[f'P{row_num+3}'] = report['summ_fire_alarm']
    ws[f'Q{row_num+3}'] = report['itog_summ_mounth']
    ws[f'C{row_num+6}'] = 'Итого охраняется:'
    ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    ws[f'D{row_num+9}'] = report['summ_senim']
    ws[f'D{row_num+8}'] = report['summ_kts']
    ws[f'C{row_num+9}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'F{row_num+6}'] = report['partners_kolvo_object']
    ws[f'C{row_num+10}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num+10}'] = report['summ_all_company']
    ws[f'C{row_num+11}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num+12}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num+12}'] = '___________________'
    ws[f'E{row_num+12}'] = 'Пак И.C.'
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=SGS_PLUS_URIK.xlsx'

    return response


@login_required
def sgs_plus_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=1, urik=False).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1


    partners_object_podkl = partners_object.objects.filter(
        company_name_id=1, urik=False
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    )



    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days)
            else:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = int(
                ((kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)
        else:
            reagirovanie = int(((kts_instance.tariff_per_mounth) / num_days_mounth) * num_days)
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services
        summ_all_company = summ_kts + summ_senim

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
            'summ_all_company': summ_all_company,
        })

    template_path = os.path.join(settings.MEDIA_ROOT, 'sgs_plus_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        # ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'F{row_num}'] = report['kts_instance'].date_podkluchenia
        # ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'G{row_num}'] = report['num_days']
        ws[f'H{row_num}'] = report['itog_rent_gsm']
        ws[f'I{row_num}'] = report['itog_telemetria']
        ws[f'J{row_num}'] = report['itog_nabludenie']
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_tehnical_services']
        ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        # ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'I{row_num+1}'] = report['summ_telemetria']
    ws[f'H{row_num+1}'] = report['summ_rent_gsm']
    ws[f'J{row_num+1}'] = report['summ_nabludenie']
    ws[f'K{row_num+1}'] = report['summ_reagirovanie']
    ws[f'L{row_num+1}'] = report['summ_tehnical_services']
    ws[f'M{row_num+1}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num+3}'] = 'Итого охраняется:'
    ws[f'C{row_num+4}'] = 'ТОО "КузетТехноСервис"'
    ws[f'F{row_num+5}'] = report['summ_senim']
    ws[f'F{row_num+4}'] = report['summ_kts']
    ws[f'C{row_num+5}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'G{row_num+3}'] = report['partners_kolvo_object']
    ws[f'C{row_num+7}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'H{row_num+7}'] = report['summ_all_company']
    ws[f'C{row_num+8}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num+9}'] = 'Исполнитель: бухгалтер'
    ws[f'F{row_num+9}'] = '___________________'
    ws[f'H{row_num+9}'] = 'Пак И.C.'
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=SGS_PLUS_FIZIKI_{now.date()}.xlsx'

    return response


# ОТЧЕТЫ АКМ ЭКСЕЛЬ ЭКСПОРТ ПО ФИЗИКАИ И ЮРИКАМ ОТДЕЛЬНО
@login_required
def reports_partners_akm(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_month = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(
        company_name_id=2
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    )
    partners_kolvo_object = partners_object.objects.filter(
        company_name_id=2
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(
        company_name_id=2, urik=True
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(
        company_name_id=2, urik=False
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (
                    kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_month - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_month - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_month
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_month
            else:
                num_days = num_days_month

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_month) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_month) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_month) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_month) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_month) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_month) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_month) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_month) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_month) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (
                                       kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_month * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(
                    reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_month) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number and kts_instance.company_name.sms_ur:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_month * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
                elif kts_instance.company_name.sms_ur:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_month * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = 0
            else:
                if kts_instance.sms_number and kts_instance.company_name.sms:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_month * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
                elif kts_instance.company_name.sms:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_month * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = 0
        else:
            itog_sms_uvedomlenie = 0

        if kts_instance.primechanie == '50% на 50%':
            summ_mounth = int((
                                          itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2)
        else:
            summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm

        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_akm.html', context)


@login_required
def akm_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=2, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=2, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    akm_fiz_count = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'akm_fiz_count': akm_fiz_count,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'akm_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['all_object_number']
        ws[f'D{row_num}'] = report['kts_instance'].name_object
        ws[f'E{row_num}'] = report['kts_instance'].adres
        ws[f'F{row_num}'] = report['kts_instance'].type_object
        ws[f'G{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'H{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'I{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['num_days']
        # ws[f'H{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_tehnical_services']
        ws[f'N{row_num}'] = report['itog_sms_uvedomlenie']
        # ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'D{row_num}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    ws[f'L{row_num}'] = report['summ_reagirovanie']
    ws[f'M{row_num}'] = report['summ_tehnical_services']
    ws[f'N{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'O{row_num}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    ws[f'E{row_num + 2}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'D{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'D{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'E{row_num + 5}'] = '________________'
    ws[f'F{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO AKM Fiziki {now.date()}.xlsx'

    return response


@login_required
def akm_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=2, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=2, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    akm_fiz_count = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'akm_fiz_count': akm_fiz_count,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'akm_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['num_days']
        ws[f'K{row_num}'] = report['itog_rent_gsm']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_nabludenie']
        ws[f'N{row_num}'] = report['reagirovanie']
        ws[f'O{row_num}'] = report['itog_tehnical_services']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['itog_fire_alarm']
        ws[f'R{row_num}'] = report['summ_mounth']
        ws[f'S{row_num}'] = report['kts_instance'].primechanie
        ws[f'T{row_num}'] = report['kts_instance'].date_otkluchenia
        row_num += 1

    ws[f'D{row_num}'] = 'Итого'
    ws[f'K{row_num}'] = report['summ_rent_gsm']
    ws[f'L{row_num}'] = report['summ_telemetria']
    ws[f'M{row_num}'] = report['summ_nabludenie']
    ws[f'N{row_num}'] = report['summ_reagirovanie']
    ws[f'O{row_num}'] = report['summ_tehnical_services']
    ws[f'P{row_num}'] = report['summ_sms_uvedomlenie']
    ws[f'Q{row_num}'] = report['summ_fire_alarm']
    ws[f'R{row_num}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    ws[f'E{row_num + 2}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'D{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'D{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'E{row_num + 5}'] = '___________________'
    ws[f'G{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO AKM Uriki {now.date()}.xlsx'

    return response

# ОТЧЕТЫ АКМ КОНЕЦ


# ОТЧЕТЫ RMG ЭКСЕЛЬ ЭКСПОРТ ПО ФИЗИКАИ И ЮРИКАМ ОТДЕЛЬНО
@login_required
def reports_partners_rmg(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=4).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=4).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=4, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=4, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth


        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days)
            else:
                itog_tehnical_services = int(
                    (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = 5000
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_rmg.html', context)


@login_required
def rmg_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=4, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    rmg_fiz_count = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=4, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        # all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = 5000
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'rmg_fiz_count': rmg_fiz_count,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'rmg_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{8}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 11
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].type_object
        ws[f'D{row_num}'] = report['kts_instance'].name_object
        ws[f'E{row_num}'] = report['kts_instance'].adres
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        # ws[f'I{row_num}'] = report['kts_instance'].date_podkluchenia
        # ws[f'K{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        # ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'J{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'K{row_num}'] = report['itog_tehnical_services']
        # ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['summ_mounth']
        ws[f'M{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'H{row_num}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    # ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    # ws[f'L{row_num}'] = report['summ_reagirovanie']
    # ws[f'M{row_num}'] = report['summ_tehnical_services']
    # ws[f'N{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'L{row_num}'] = report['itog_summ_mounth']
    # ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    # ws[f'G{row_num + 2}'] = report['rmg_fiz_count']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'D{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'E{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'D{row_num + 5}'] = 'Сверку проверили:'
    ws[f'C{row_num + 6}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'E{row_num + 6}'] = '___________________/'
    ws[f'G{row_num + 6}'] = 'Пак И.C./'
    ws[f'C{row_num + 7}'] = 'ТОО "RMG GROUP"'
    ws[f'E{row_num + 7}'] = '____________________/'
    ws[f'G{row_num + 7}'] = '__________________/'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO RMG Fiziki {now.date()}.xlsx'

    return response


@login_required
def rmg_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month-1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=4, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    rmg_fiz_count = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=4, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        # all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = 5000
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.sms_number:
                itog_sms_uvedomlenie = int(
                    ((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth) * num_days)
            else:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms / num_days_mounth) * num_days)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'rmg_fiz_count': rmg_fiz_count,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'rmg_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[
            f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].type_object
        ws[f'D{row_num}'] = report['kts_instance'].name_object
        ws[f'E{row_num}'] = report['kts_instance'].adres
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        # ws[f'I{row_num}'] = report['kts_instance'].date_podkluchenia
        # ws[f'K{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        # ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'J{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'K{row_num}'] = report['itog_tehnical_services']
        # ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['summ_mounth']
        ws[f'M{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'H{row_num}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    # ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    # ws[f'L{row_num}'] = report['summ_reagirovanie']
    # ws[f'M{row_num}'] = report['summ_tehnical_services']
    # ws[f'N{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'L{row_num}'] = report['itog_summ_mounth']
    # ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    # ws[f'G{row_num + 2}'] = report['rmg_fiz_count']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'B{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'E{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 5}'] = 'Сверку проверили:'
    ws[f'C{row_num + 6}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'E{row_num + 6}'] = '___________________/'
    ws[f'G{row_num + 6}'] = 'Пак И.C.'
    ws[f'C{row_num + 7}'] = 'ТОО "RMG GROUP"'
    ws[f'E{row_num + 7}'] = '____________________/'
    ws[f'G{row_num + 7}'] = '__________________/'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO RMG Uriki {now.date()}.xlsx'

    return response
# ОТЧЕТЫ АКМ КОНЕЦ








# ОТЧЕТЫ kaz-kuzet ЭКСЕЛЬ ЭКСПОРТ ПО ФИЗИКАИ И ЮРИКАМ ОТДЕЛЬНО
@login_required
def reports_partners_kazkuzet(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(
        company_name_id=3
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    )
    partners_kolvo_object = partners_object.objects.filter(
        company_name_id=3
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(
        company_name_id=3, urik=True
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(
        company_name_id=3, urik=False
    ).exclude(
        Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))
    ).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_kazkuzet.html', context)


@login_required
def kazkuzet_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=3, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=3, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    kazkuzet = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'kazkuzet': kazkuzet,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'kazkuzet_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['num_days']
        # ws[f'H{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_tehnical_services']
        ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        # ws[f'P{row_num}'] = report['itog_fire_alarm']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'D{row_num}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    # ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    ws[f'K{row_num}'] = report['summ_reagirovanie']
    ws[f'L{row_num}'] = report['summ_tehnical_services']
    ws[f'M{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'N{row_num}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    ws[f'E{row_num + 2}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'D{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'D{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'E{row_num + 5}'] = '___________________'
    ws[f'G{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO KazKuzet Fiziki {now.date()}.xlsx'

    return response



@login_required
def kazkuzet_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=3, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=3, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    kazkuzet = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'kazkuzet': kazkuzet,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'kazkuzet_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['num_days']
        # ws[f'H{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_tehnical_services']
        ws[f'M{row_num}'] = report['itog_fire_alarm']
        ws[f'N{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'D{row_num}'] = 'Итого'
    ws[f'K{row_num}'] = report['summ_reagirovanie']
    ws[f'L{row_num}'] = report['summ_tehnical_services']
    ws[f'M{row_num}'] = report['summ_fire_alarm']
    ws[f'N{row_num}'] = report['summ_sms_uvedomlenie']
    ws[f'O{row_num}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 2}'] = 'Итого охраняется:'
    ws[f'E{row_num + 2}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'D{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'D{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'D{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'E{row_num + 5}'] = '___________________'
    ws[f'G{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO KazKuzet Uriki {now.date()}.xlsx'

    return response

# ОТЧЕТЫ kaz-kuzet КОНЕЦ



# ОТЧЕТЫ SGS начало
@login_required
def reports_partners_sgs(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=5).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=5).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=5, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=5, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_sgs.html', context)


@login_required
def sgs_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=5, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=5, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'sgs_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        # ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'B{row_num}'] = report['kts_instance'].name_object
        ws[f'C{row_num}'] = report['kts_instance'].adres
        ws[f'D{row_num}'] = report['kts_instance'].type_object
        ws[f'E{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'I{row_num}'] = report['itog_fire_alarm']
        ws[f'J{row_num}'] = report['num_days']
        # ws[f'H{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        # ws[f'K{row_num}'] = report['reagirovanie']
        # ws[f'L{row_num}'] = report['itog_tehnical_services']
        # ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'K{row_num}'] = report['summ_mounth']
        ws[f'L{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'B{row_num+1}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    # ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    # ws[f'K{row_num}'] = report['summ_reagirovanie']
    # ws[f'L{row_num}'] = report['summ_tehnical_services']
    # ws[f'M{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'K{row_num+1}'] = report['itog_summ_mounth']
    ws[f'B{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'B{row_num + 5}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 5}'] = report['itog_summ_mounth']
    ws[f'B{row_num + 6}'] = '(В том числе НДС 12%)'
    ws[f'B{row_num + 7}'] = 'Бухглалтер ТОО "System of Global Safety" '
    ws[f'D{row_num + 7}'] = '_________________________'
    ws[f'B{row_num + 8}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 8}'] = '___________________'
    ws[f'F{row_num + 8}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "SGS" Fiziki {now.date()}.xlsx'

    return response


@login_required
def sgs_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=5, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=5, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'sgs_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        # ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'B{row_num}'] = report['kts_instance'].name_object
        ws[f'C{row_num}'] = report['kts_instance'].adres
        ws[f'D{row_num}'] = report['kts_instance'].type_object
        ws[f'E{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'H{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'I{row_num}'] = report['itog_fire_alarm']
        ws[f'J{row_num}'] = report['num_days']
        # ws[f'H{row_num}'] = report['itog_rent_gsm']
        # ws[f'I{row_num}'] = report['itog_telemetria']
        # ws[f'J{row_num}'] = report['itog_nabludenie']
        # ws[f'K{row_num}'] = report['reagirovanie']
        # ws[f'L{row_num}'] = report['itog_tehnical_services']
        # ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'K{row_num}'] = report['summ_mounth']
        ws[f'L{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'B{row_num+2}'] = 'Итого'
    # ws[f'I{row_num + 1}'] = report['summ_telemetria']
    # ws[f'H{row_num + 1}'] = report['summ_rent_gsm']
    # ws[f'J{row_num}'] = report['summ_tariff_per_mounth']
    # ws[f'K{row_num}'] = report['summ_reagirovanie']
    # ws[f'L{row_num}'] = report['summ_tehnical_services']
    # ws[f'M{row_num}'] = report['summ_sms_uvedomlenie']
    # ws[f'P{row_num+1}'] = report['summ_fire_alarm']
    ws[f'K{row_num+2}'] = report['itog_summ_mounth']
    ws[f'B{row_num + 4}'] = 'Итого охраняется:'
    ws[f'D{row_num + 4}'] = report['partners_kolvo_object']
    # ws[f'C{row_num+8}'] = 'ТОО "КузетТехноСервис"'
    # ws[f'D{row_num+9}'] = report['summ_senim']
    # ws[f'D{row_num+8}'] = report['summ_kts']
    # ws[f'C{row_num+9}'] = 'ТОО "КузетСенiм"'
    ws[f'B{row_num + 5}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 5}'] = report['itog_summ_mounth']
    ws[f'B{row_num + 6}'] = '(В том числе НДС 12%)'
    ws[f'B{row_num + 7}'] = 'Бухглалтер ТОО "System of Global Safety" '
    ws[f'D{row_num + 7}'] = '_________________________'
    ws[f'B{row_num + 8}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 8}'] = '_________________________'
    ws[f'G{row_num + 8}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "SGS" Uriki {now.date()}.xlsx'

    return response

# ОТЧЕТЫ SGS конец




@login_required
def reports_partners_ipkim(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=6).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=6).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=6, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=6, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_ipkim.html', context)



@login_required
def ipkim_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=6, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=6, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'ipkim_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физицеским лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['num_days']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'H{row_num + 6}'] = '_________________'
    ws[f'K{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "KIM" Fiziki {now.date()}.xlsx'

    return response



@login_required
def ipkim_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=6, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=6, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'ipkim_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['num_days']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['sgs']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'F{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "KIM" Urik {now.date()}.xlsx'

    return response


@login_required
def reports_partners_kuzets(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=7).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=7).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=7, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=7, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
        'summ_kts': summ_kts,
        'summ_senim': summ_senim,
    }

    return render(request, 'dogovornoy/reports_partners_kuzets.html', context)



@login_required
def kuzets_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=7, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=7, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'kuzets_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'F{row_num}'] = report['num_days']
        ws[f'G{row_num}'] = report['itog_rent_gsm']
        ws[f'H{row_num}'] = report['itog_nabludenie']
        ws[f'I{row_num}'] = report['reagirovanie']
        ws[f'J{row_num}'] = report['itog_tehnical_services']
        ws[f'K{row_num}'] = report['itog_telemetria']
        ws[f'L{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'M{row_num}'] = report['summ_mounth']
        ws[f'N{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+3}'] = 'Итого'
    ws[f'G{row_num + 3}'] = report['summ_rent_gsm']
    ws[f'H{row_num + 3}'] = report['summ_nabludenie']
    ws[f'I{row_num + 3}'] = report['summ_reagirovanie']
    ws[f'J{row_num + 3}'] = report['summ_tehnical_services']
    ws[f'K{row_num + 3}'] = report['summ_telemetria']
    ws[f'L{row_num + 3}'] = report['summ_sms_uvedomlenie']
    ws[f'M{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 7}'] = 'Итого охраняется:'
    ws[f'D{row_num + 7}'] = report['sgs']
    ws[f'C{row_num + 8}'] = 'ТОО "КузетТехноСервис"'
    ws[f'D{row_num + 8}'] = report['summ_kts']
    ws[f'C{row_num + 9}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'D{row_num + 9}'] = report['summ_senim']
    ws[f'C{row_num + 10}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 10}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 11}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 12}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 12}'] = '_________________'
    ws[f'E{row_num + 12}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "Kuzet-S" Fiziki {now.date()}.xlsx'

    return response



@login_required
def kuzets_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=7, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=7, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services
        summ_all_company = summ_kts + summ_senim

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
            'summ_all_company': summ_all_company,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'kuzets_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        # ws[f'C{row_num}'] = report['all_object_number']
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'F{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'G{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'H{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['itog_telemetria']
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['itog_nabludenie']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_tehnical_services']
        ws[f'N{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+4}'] = 'Итого'
    ws[f'I{row_num+4}'] = report['summ_telemetria']
    ws[f'J{row_num+4}'] = report['summ_rent_gsm']
    ws[f'K{row_num+4}'] = report['summ_nabludenie']
    ws[f'L{row_num+4}'] = report['summ_reagirovanie']
    ws[f'M{row_num+4}'] = report['summ_tehnical_services']
    ws[f'N{row_num+4}'] = report['summ_sms_uvedomlenie']
    ws[f'O{row_num+4}'] = report['summ_all_company']
    ws[f'C{row_num + 7}'] = 'Итого охраняется:'
    ws[f'D{row_num + 7}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 8}'] = 'ТОО "КузетТехноСервис"'
    ws[f'D{row_num + 8}'] = report['summ_kts']
    ws[f'C{row_num + 9}'] = 'ТОО "Кузет-Сенiм"'
    ws[f'D{row_num + 9}'] = report['summ_senim']
    ws[f'C{row_num + 10}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 10}'] = report['summ_all_company']
    ws[f'C{row_num + 11}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 12}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 12}'] = '_________________'
    ws[f'E{row_num + 12}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "Kuzet-S" Uriki {now.date()}.xlsx'

    return response



@login_required
def reports_partners_samohvalov(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=8).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=8).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=8, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=8, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        if kts_instance.primechanie == '50% на 50%':
            summ_mounth = int((itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2)
        else:
            summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm

        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_samohvalov.html', context)


@login_required
def samohvalov_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=8, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=8, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        if kts_instance.primechanie == '50% на 50%':
            summ_mounth = (itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2
        else:
            summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm

        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'samohvalov_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_fire_alarm']
        ws[f'K{row_num}'] = report['itog_tehnical_services']
        ws[f'L{row_num}'] = report['num_days']
        ws[f'M{row_num}'] = report['summ_mounth']
        ws[f'N{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num}'] = 'Итого'
    ws[f'M{row_num}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 2}'] = 'Итого охраняется:'
    ws[f'E{row_num + 2}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 5}'] = '_________________'
    ws[f'E{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "Samohvalov" Fiziki {now.date()}.xlsx'

    return response



@login_required
def samohvalov_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=8, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=8, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0
    itog_tehnical_services = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        if kts_instance.primechanie == '50% на 50%':
            summ_mounth = (itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2
        else:
            summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm

        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth
        summ_kts = summ_telemetria + summ_rent_gsm + summ_nabludenie + summ_sms_uvedomlenie
        summ_senim = summ_reagirovanie + summ_tehnical_services

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
            'summ_kts': summ_kts,
            'summ_senim': summ_senim,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'samohvalov_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[
            f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_fire_alarm']
        ws[f'K{row_num}'] = report['itog_tehnical_services']
        ws[f'L{row_num}'] = report['num_days']
        ws[f'M{row_num}'] = report['summ_mounth']
        ws[f'N{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+2}'] = 'Итого'
    ws[f'M{row_num+2}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = 'Итого охраняется:'
    ws[f'D{row_num + 5}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 6}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'E{row_num + 6}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 7}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 8}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 8}'] = '_________________'
    ws[f'E{row_num + 8}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ip "Samohvalov" uriki {now.date()}.xlsx'

    return response







@login_required
def reports_partners_sobsecutity(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=9).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=9).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=9, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=9, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_sobsecurity.html', context)




@login_required
def sobsecutity_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=9, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=9, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'sobsecutity_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'M{row_num}'] = report['num_days']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num}'] = 'Итого'
    ws[f'N{row_num}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 2}'] = 'Итого охраняется:'
    ws[f'D{row_num + 2}'] = report['sgs']
    ws[f'C{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 5}'] = '_________________'
    ws[f'F{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "SOB Security" Fiziki {now.date()}.xlsx'

    return response



@login_required
def sobsecutity_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=9, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=9, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'sobsecutity_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'M{row_num}'] = report['num_days']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num}'] = 'Итого'
    ws[f'N{row_num}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['sgs']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'F{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "SOB Security" Urik {now.date()}.xlsx'

    return response


@login_required
def reports_partners_egida(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=15).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=15).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=15, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=15, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_egida.html', context)


@login_required
def egida_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=15, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=15, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'egida_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['summ_mounth']
        ws[f'M{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'L{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Egida Druzhina" Fiziki {now.date()}.xlsx'

    return response


@login_required
def egida_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=15, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=15, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'egida_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['summ_mounth']
        ws[f'M{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'L{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Egida Druzhina" Uriki {now.date()}.xlsx'

    return response








@login_required
def reports_partners_eyewatch(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=10).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=10).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=10, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=10, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_eyewatch.html', context)


@login_required
def eyewatch_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=10, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object_podkl.count()

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=10, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'eyewatch_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'
    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'Q{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['sgs']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Eye Watch" Fiziki {now.date()}.xlsx'

    return response


@login_required
def eyewatch_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=10, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=10, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=10, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
                        'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'eyewatch_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num}'] = 'Итого'
    ws[f'Q{row_num}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Eye Watch" Uriki {now.date()}.xlsx'

    return response




@login_required
def reports_partners_iviscom(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=11).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=11).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=11, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=11, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_iviscom.html', context)


@login_required
def iviscom_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=11, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=11, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=11, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'iviscom_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num + 1}'] = 'Итого'
    ws[f'N{row_num + 1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ТОО "IVISCOM" Fiziki {now.date()}.xlsx'

    return response




@login_required
def iviscom_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=11, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=11, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=11, urik=True).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'iviscom_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "IVISCOM" Uriki {now.date()}.xlsx'

    return response




@login_required
def reports_partners_eurasian(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=12).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=12).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=12, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=12, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_eurasian.html', context)


@login_required
def eurasian_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=12, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=12, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=12, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'eurasian_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num + 1}'] = 'Итого'
    ws[f'Q{row_num + 1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "EurasianSecuritySystem" Fiziki {now.date()}.xlsx'

    return response




@login_required
def eurasian_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=12, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=12, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=12, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'eurasian_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'Q{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '__________________________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "EurasianSecuritySystem" Uriki {now.date()}.xlsx'

    return response





@login_required
def reports_partners_bmkz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=13).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=13).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=13, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=13, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_bmkz.html', context)


@login_required
def bmkz_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=1, urik=True).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=13, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=13, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'bmkz_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'N{row_num}'] = report['itog_tehnical_services']
        ws[f'O{row_num}'] = report['itog_fire_alarm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    # ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    # ws[f'D{row_num + 3}'] = report['sgs']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    # ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'H{row_num + 6}'] = '_________________'
    ws[f'K{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "B.M.kz Security" Fiziki {now.date()}.xlsx'

    return response


@login_required
def bmkz_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=13, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=13, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=13, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'bmkz_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{6}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 9
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_telemetria']
        ws[f'M{row_num}'] = report['itog_rent_gsm']
        ws[f'P{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'Q{row_num}'] = report['summ_mounth']
        ws[f'R{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num}'] = 'Итого'
    ws[f'Q{row_num}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 2}'] = 'Итого охраняется:'
    ws[f'D{row_num + 2}'] = report['sgs']
    ws[f'C{row_num + 3}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 3}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 4}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 5}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 5}'] = '_________________'
    ws[f'E{row_num + 5}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "BMkz Security" Uriki {now.date()}.xlsx'

    return response







@login_required
def reports_partners_monolit(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=14).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=14).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=14, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=14, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_monolit.html', context)


@login_required
def monolit_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=14, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=14, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=14, urik=False).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'monolit_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[
            f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        # ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'H{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['reagirovanie']
        # ws[f'K{row_num}'] = report['itog_telemetria']
        ws[f'K{row_num}'] = report['itog_rent_gsm']
        ws[f'L{row_num}'] = report['itog_tehnical_services']
        ws[f'M{row_num}'] = report['itog_fire_alarm']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num + 1}'] = 'Итого'
    ws[f'N{row_num + 1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Monolite Security" Fiziki {now.date()}.xlsx'

    return response


@login_required
def monolit_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=14, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=14, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=14, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'monolit_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        # ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'F{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'G{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'H{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['reagirovanie']
        # ws[f'K{row_num}'] = report['itog_telemetria']
        ws[f'K{row_num}'] = report['itog_rent_gsm']
        ws[f'L{row_num}'] = report['itog_tehnical_services']
        ws[f'M{row_num}'] = report['itog_fire_alarm']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Monolite Security" Uriki {now.date()}.xlsx'

    return response



@login_required
def reports_kolvo(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month-1, 1, tzinfo=timezone.utc)
    end_of_month = timezone.datetime(now.year, now.month-1, calendar.monthrange(now.year, now.month-1)[1], tzinfo=timezone.utc)
    next_start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc).date()
    next_end_of_month = timezone.datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1], tzinfo=timezone.utc).date()

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)


    companies = rekvizity.objects.all()

    reports = []

    for company in companies:

        # 2 Все объекты до выбранной даты
        kts_podkl = kts.objects.filter(company_name_id=company.id).exclude(
            Q(date_otklulchenia__lt=start_of_month) & Q(date_podkluchenia__lte=F('date_otklulchenia')))


        # Всего на начало выбранного месяца
        kts_count_podkl = kts_podkl.filter(
                                            date_podkluchenia__lte=start_of_month,
                                            urik=True,
                                            exclude_from_report=False).count()


        kts_fiz_podkl = kts_podkl.filter(
                                            date_podkluchenia__lte=start_of_month,
                                            urik=False,
                                            exclude_from_report=False).count()


        # принято(в т.ч.после вр.снятия )
        kolvo_podkl_obj = kts_podkl.filter(urik=True,
                                            date_podkluchenia__gte=start_of_month,
                                            date_podkluchenia__lte=end_of_month,
                                           exclude_from_report=False).count()

        kolvo_podkl_fiz = kts_podkl.filter(urik=False,
                                             date_podkluchenia__gte=start_of_month,
                                             date_podkluchenia__lte=end_of_month,
                                           exclude_from_report=False).count()

        # расторженно (в т.ч.после вр.снятия )
        kolvo_otkl_obj = kts_podkl.filter(urik=True,
                                            date_otklulchenia__gte=start_of_month,
                                            date_otklulchenia__lte=end_of_month,
                                          exclude_from_report=False).count()

        kolvo_otkl_fiz = kts_podkl.filter(urik=False,
                                            date_otklulchenia__gte=start_of_month,
                                            date_otklulchenia__lte=end_of_month,
                                          exclude_from_report=False).count()


        # Всего на конец выбранного месяца
        kts_count_podkl_end = (kts_count_podkl + kolvo_podkl_obj) - kolvo_otkl_obj

        kts_fiz_podkl_end = (kts_fiz_podkl + kolvo_podkl_fiz) - kolvo_otkl_fiz


        # экипажи физические лица
        gruppa_reagirovania_911_fiz = kts_podkl.filter(urik=False,
                                                         gruppa_reagirovania='911', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_bravo21_fiz = kts_podkl.filter(urik=False, gruppa_reagirovania='Браво-21', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_sms_fiz = kts_podkl.filter(urik=False, gruppa_reagirovania='СМС', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_asker_fiz = kts_podkl.filter(urik=False, gruppa_reagirovania='Эскер', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_zardem_fiz = kts_podkl.filter(urik=False, gruppa_reagirovania='Жардем', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_kuguar_fiz = kts_podkl.filter(urik=False, gruppa_reagirovania='Кугуар', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()


        kolvo_ekipazh_fiz = kts_fiz_podkl_end - (gruppa_reagirovania_911_fiz + gruppa_reagirovania_bravo21_fiz + gruppa_reagirovania_sms_fiz +
                             gruppa_reagirovania_asker_fiz + gruppa_reagirovania_zardem_fiz + gruppa_reagirovania_kuguar_fiz)

        # экипажи юридические лица
        gruppa_reagirovania_911_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='911', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_bravo21_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='Браво-21', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_sms_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='СМС', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_asker_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='Эскер', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_zardem_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='Жардем', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()

        gruppa_reagirovania_kuguar_ur = kts_podkl.filter(urik=True, gruppa_reagirovania='Кугуар', exclude_from_report=False).exclude(date_otklulchenia__lte=end_of_month).count()


        kolvo_ekipazh_ur = kts_count_podkl_end - (gruppa_reagirovania_911_ur + gruppa_reagirovania_bravo21_ur + gruppa_reagirovania_sms_ur
                                                  + gruppa_reagirovania_asker_ur + gruppa_reagirovania_zardem_ur + gruppa_reagirovania_kuguar_ur)


        reports.append({
            'companies': companies,
            'kts_podkl': kts_podkl,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'kts_count_podkl_end': kts_count_podkl_end,
            'kts_fiz_podkl_end': kts_fiz_podkl_end,
            'start_of_month': start_of_month,
            'end_of_month': end_of_month,
            'kolvo_podkl_obj': kolvo_podkl_obj,
            'kolvo_podkl_fiz': kolvo_podkl_fiz,
            'kolvo_otkl_obj': kolvo_otkl_obj,
            'kolvo_otkl_fiz': kolvo_otkl_fiz,
            'gruppa_reagirovania_911_fiz': gruppa_reagirovania_911_fiz,
            'gruppa_reagirovania_sms_fiz': gruppa_reagirovania_sms_fiz,
            'gruppa_reagirovania_asker_fiz': gruppa_reagirovania_asker_fiz,
            'gruppa_reagirovania_zardem_fiz': gruppa_reagirovania_zardem_fiz,
            'gruppa_reagirovania_bravo21_fiz': gruppa_reagirovania_bravo21_fiz,
            'gruppa_reagirovania_911_ur': gruppa_reagirovania_911_ur,
            'gruppa_reagirovania_sms_ur': gruppa_reagirovania_sms_ur,
            'gruppa_reagirovania_asker_ur': gruppa_reagirovania_asker_ur,
            'gruppa_reagirovania_zardem_ur': gruppa_reagirovania_zardem_ur,
            'gruppa_reagirovania_bravo21_ur': gruppa_reagirovania_bravo21_ur,
            'gruppa_reagirovania_kuguar_ur': gruppa_reagirovania_kuguar_ur,
            'gruppa_reagirovania_kuguar_fiz': gruppa_reagirovania_kuguar_fiz,
            'kolvo_ekipazh_fiz': kolvo_ekipazh_fiz,
            'kolvo_ekipazh_ur': kolvo_ekipazh_ur,
        })
        context = {'reports': reports, 'start_of_month': start_of_month, 'end_of_month': end_of_month}
    return render(request, 'dogovornoy/reports_kolvo.html', context)




# Вычисляет итоговую сумму по каждому клиенту партнера
def calculate_monthly_sum(kts_instance, start_of_month, end_of_month, num_days_month):
    if kts_instance.date_otkluchenia:
        if isinstance(start_of_month, datetime):
            start_of_month = start_of_month.date()
        if isinstance(end_of_month, datetime):
            end_of_month = end_of_month.date()

        if (kts_instance.date_otkluchenia > start_of_month) and (
                kts_instance.date_podkluchenia < start_of_month):
            num_days = (kts_instance.date_otkluchenia - start_of_month).days
        elif (kts_instance.date_otkluchenia >= start_of_month) and (
                kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
            num_days = num_days_month - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
        elif kts_instance.date_otkluchenia > start_of_month:
            num_days = num_days_month - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
        elif kts_instance.date_otkluchenia < start_of_month:
            num_days = num_days_month
        else:
            num_days = (kts_instance.date_otkluchenia - start_of_month).days
    else:
        if kts_instance.date_podkluchenia:
            if kts_instance.date_podkluchenia > start_of_month:
                num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
            else:
                num_days = num_days_month
        else:
            num_days = num_days_month

    if kts_instance.telemetria:
        itog_telemetria = int((kts_instance.company_name.telemetria / num_days_month) * num_days)
    else:
        itog_telemetria = 0

    if kts_instance.rent_gsm:
        if kts_instance.urik:
            itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_month) * num_days)
        else:
            itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_month) * num_days)
    else:
        itog_rent_gsm = 0

    if kts_instance.nabludenie:
        if kts_instance.urik:
            itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_month) * num_days)
        else:
            itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_month) * num_days)
    else:
        itog_nabludenie = 0

    if kts_instance.tehnical_services:
        if kts_instance.urik:
            itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_month) * num_days
            itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                itog_tehnical_services) > 0.5 else math.floor(itog_tehnical_services)
        else:
            itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_month) * num_days
            itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                itog_tehnical_services) > 0.5 else math.floor(itog_tehnical_services)
    else:
        itog_tehnical_services = 0

    if kts_instance.fire_alarm:
        if kts_instance.urik:
            itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_month) * num_days)
        else:
            itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_month) * num_days)
    else:
        itog_fire_alarm = 0

    if kts_instance.urik:
        if kts_instance.tariff_per_mounth > 30:
            reagirovanie = kts_instance.tariff_per_mounth
        else:
            reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_month * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)
    else:
        reagirovanie = (kts_instance.tariff_per_mounth / num_days_month) * num_days
        reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
            reagirovanie)

    if kts_instance.sms_uvedomlenie:
        if kts_instance.urik:
            if kts_instance.sms_number and kts_instance.company_name.sms_ur:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_month * num_days)
                itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
            elif kts_instance.company_name.sms_ur:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_month * num_days)
                itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
            else:
                itog_sms_uvedomlenie = 0
        else:
            if kts_instance.sms_number and kts_instance.company_name.sms:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_month * num_days)
                itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
            elif kts_instance.company_name.sms:
                itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_month * num_days)
                itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(itog_sms_uvedomlenie) > 0.5 else math.floor(itog_sms_uvedomlenie)
            else:
                itog_sms_uvedomlenie = 0
    else:
        itog_sms_uvedomlenie = 0

    if kts_instance.primechanie == '50% на 50%':
        summ_mounth = int((itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm) / 2)
    else:
        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm

    return summ_mounth



@login_required
def partner_reports_kolvo(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1], tzinfo=timezone.utc).date()

    next_month = now + timedelta(days=calendar.monthrange(now.year, now.month-1)[1])
    next_start_of_month = datetime(next_month.year, next_month.month, 1, tzinfo=timezone.utc).date()
    next_end_of_month = datetime(next_month.year, next_month.month,
                                 calendar.monthrange(next_month.year, next_month.month)[1], tzinfo=timezone.utc).date()

    num_days_month = calendar.monthrange(now.year, now.month - 1)[1]

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_of_month = datetime.strptime(end_date, "%Y-%m-%d").date()

    partners = partners_rekvizity.objects.all()
    reports = []
    itog_all = 0
    kolvo_podkl_obj_summ = 0
    kolvo_otkl_obj_summ = 0
    kts_itog_object_all = 0
    kts_itog_object_end_all = 0
    kts_fiz_podkl_all = 0
    kts_fiz_podkl_end_all = 0
    kts_count_podkl_all = 0
    kts_count_podkl_end_all = 0

    for partner in partners:
        connected_objects = partners_object.objects.filter(company_name=partner).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))

        kts_podkl_start = connected_objects.filter(
            Q(date_otkluchenia__gte=start_of_month, date_otkluchenia__lte=next_end_of_month) |
            Q(date_otkluchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
        )

        kts_count_podkl = connected_objects.filter(
            Q(date_otkluchenia__gte=start_of_month, date_otkluchenia__lte=next_end_of_month) |
            Q(date_otkluchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
            urik=True,
        ).aggregate(count=Count('id'))['count']

        kts_count_podkl_all += kts_count_podkl

        kts_fiz_podkl = connected_objects.filter(
            Q(date_otkluchenia__gte=start_of_month, date_otkluchenia__lte=next_end_of_month) |
            Q(date_otkluchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
            urik=False,
        ).aggregate(count=Count('id'))['count']

        kts_fiz_podkl_all += kts_fiz_podkl

        kts_itog_object = kts_count_podkl + kts_fiz_podkl
        kts_itog_object_all  += kts_itog_object

        kts_count_podkl_end = connected_objects.filter(
            Q(date_otkluchenia__gte=next_start_of_month, date_otkluchenia__lte=next_end_of_month) |
            Q(date_otkluchenia__isnull=True),
            date_podkluchenia__lte=end_of_month,
            urik=True,
        ).aggregate(count=Count('id'))['count']

        kts_count_podkl_end_all += kts_count_podkl_end

        kts_fiz_podkl_end = connected_objects.filter(
            Q(date_otkluchenia__gte=next_start_of_month, date_otkluchenia__lte=next_end_of_month) |
            Q(date_otkluchenia__isnull=True),
            date_podkluchenia__lte=end_of_month,
            urik=False,
        ).aggregate(count=Count('id'))['count']

        kts_fiz_podkl_end_all += kts_fiz_podkl_end

        kts_itog_object_end = kts_count_podkl_end + kts_fiz_podkl_end
        kts_itog_object_end_all += kts_itog_object_end

        kolvo_podkl_obj = connected_objects.filter(
            date_podkluchenia__gte=start_of_month,
            date_podkluchenia__lte=end_of_month
        ).aggregate(count=Count('id'))['count']

        kolvo_podkl_obj_summ += kolvo_podkl_obj

        kolvo_otkl_obj = connected_objects.filter(
            date_otkluchenia__gte=start_of_month,
            date_otkluchenia__lte=end_of_month,
        ).aggregate(count=Count('id'))['count']

        kolvo_otkl_obj_summ += kolvo_otkl_obj

        podkl_otlk_raznica = kolvo_podkl_obj - kolvo_otkl_obj

        rost_all = kolvo_podkl_obj_summ - kolvo_otkl_obj_summ

        itog_summ_mounth = 0
        itog_summ_mounth_start = 0

        for kts_instance in connected_objects:
            summ_mounth = calculate_monthly_sum(kts_instance, start_of_month, end_of_month, num_days_month)
            itog_summ_mounth += summ_mounth
            itog_all += summ_mounth

        for kts_instance in kts_podkl_start:
            summ_mounth_start = calculate_monthly_sum(kts_instance, start_of_month, end_of_month, num_days_month)
            itog_summ_mounth_start += summ_mounth_start

        money_raznica = itog_summ_mounth - itog_summ_mounth_start

        reports.append({
            'partner': partner,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'kts_count_podkl_end': kts_count_podkl_end,
            'kts_fiz_podkl_end': kts_fiz_podkl_end,
            'kts_itog_object': kts_itog_object,
            'kts_itog_object_end': kts_itog_object_end,
            'kolvo_podkl_obj': kolvo_podkl_obj,
            'kolvo_otkl_obj': kolvo_otkl_obj,
            'podkl_otlk_raznica': podkl_otlk_raznica,
            'itog_summ_mounth': itog_summ_mounth,
            'itog_summ_mounth_start': itog_summ_mounth_start,
            'money_raznica': money_raznica,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'itog_all': itog_all,
        'kolvo_podkl_obj_summ': kolvo_podkl_obj_summ,
        'kolvo_otkl_obj_summ': kolvo_otkl_obj_summ,
        'rost_all': rost_all,
        'kts_itog_object_end_all': kts_itog_object_end_all,
        'kts_itog_object_all': kts_itog_object_all,
        'kts_fiz_podkl_all': kts_fiz_podkl_all,
        'kts_fiz_podkl_end_all': kts_fiz_podkl_end_all,
        'kts_count_podkl_all': kts_count_podkl_all,
        'kts_count_podkl_end_all': kts_count_podkl_end_all,
    }

    return render(request, 'dogovornoy/partner_reports.html', context)



@login_required
def kts_reports_kolvo(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1], tzinfo=timezone.utc).date()

    next_month = now + timedelta(days=calendar.monthrange(now.year, now.month-1)[1])
    next_start_of_month = datetime(next_month.year, next_month.month, 1, tzinfo=timezone.utc).date()
    next_end_of_month = datetime(next_month.year, next_month.month,
                                 calendar.monthrange(next_month.year, next_month.month)[1], tzinfo=timezone.utc).date()

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_of_month = datetime.strptime(end_date, "%Y-%m-%d").date()

    partners = rekvizity.objects.all()
    reports = []
    itog_all_start = 0
    kolvo_podkl_obj_summ = 0
    kolvo_otkl_obj_summ = 0
    kts_itog_object_all = 0
    kts_itog_object_end_all = 0
    kts_fiz_podkl_all = 0
    kts_fiz_podkl_end_all = 0
    kts_count_podkl_all = 0
    kts_count_podkl_end_all = 0
    itog_all_end = 0

    for partner in partners:
        connected_objects = kts.objects.filter(company_name=partner)

        kts_money_start = connected_objects.filter(
            Q(date_otklulchenia__gte=start_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
        )

        kts_money_end = connected_objects.filter(
            Q(date_otklulchenia__gte=end_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lt=end_of_month,
        )

        kts_count_podkl = connected_objects.filter(
            Q(date_otklulchenia__gte=start_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
            urik=True,
        ).aggregate(count=Count('id'))['count']

        kts_count_podkl_all += kts_count_podkl

        kts_fiz_podkl = connected_objects.filter(
            Q(date_otklulchenia__gte=start_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lt=start_of_month,
            urik=False,
        ).aggregate(count=Count('id'))['count']

        kts_fiz_podkl_all += kts_fiz_podkl

        kts_itog_object = kts_count_podkl + kts_fiz_podkl
        kts_itog_object_all  += kts_itog_object

        kts_count_podkl_end = connected_objects.filter(
            Q(date_otklulchenia__gte=end_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lte=end_of_month,
            urik=True,
        ).aggregate(count=Count('id'))['count']

        kts_count_podkl_end_all += kts_count_podkl_end

        kts_fiz_podkl_end = connected_objects.filter(
            Q(date_otklulchenia__gte=end_of_month) |
            Q(date_otklulchenia__isnull=True),
            date_podkluchenia__lte=end_of_month,
            urik=False,
        ).aggregate(count=Count('id'))['count']

        kts_fiz_podkl_end_all += kts_fiz_podkl_end

        kts_itog_object_end = kts_count_podkl_end + kts_fiz_podkl_end
        kts_itog_object_end_all += kts_itog_object_end

        kolvo_podkl_obj = connected_objects.filter(
            date_podkluchenia__gte=start_of_month,
            date_podkluchenia__lte=end_of_month
        ).aggregate(count=Count('id'))['count']

        kolvo_podkl_obj_summ += kolvo_podkl_obj

        kolvo_otkl_obj = connected_objects.filter(
            date_otklulchenia__gte=start_of_month,
            date_otklulchenia__lte=end_of_month,
        ).aggregate(count=Count('id'))['count']

        kolvo_otkl_obj_summ += kolvo_otkl_obj

        podkl_otlk_raznica = kolvo_podkl_obj - kolvo_otkl_obj

        rost_all = kolvo_podkl_obj_summ - kolvo_otkl_obj_summ

        itog_summ_mounth_end = 0
        itog_summ_mounth_start = 0

        for kts_instance in kts_money_start:
            if kts_instance.abon_plata:
                itog_summ_mounth_start += kts_instance.abon_plata
                itog_all_start += kts_instance.abon_plata

        for kts_instance in kts_money_end:
            if kts_instance.abon_plata:
                itog_summ_mounth_end += kts_instance.abon_plata
                itog_all_end += kts_instance.abon_plata

        money_raznica = itog_summ_mounth_end - itog_summ_mounth_start
        money_raznica_all = itog_all_end - itog_all_start

        reports.append({
            'partner': partner,
            'kts_count_podkl': kts_count_podkl,
            'kts_fiz_podkl': kts_fiz_podkl,
            'kts_count_podkl_end': kts_count_podkl_end,
            'kts_fiz_podkl_end': kts_fiz_podkl_end,
            'kts_itog_object': kts_itog_object,
            'kts_itog_object_end': kts_itog_object_end,
            'kolvo_podkl_obj': kolvo_podkl_obj,
            'kolvo_otkl_obj': kolvo_otkl_obj,
            'podkl_otlk_raznica': podkl_otlk_raznica,
            'itog_summ_mounth_end': itog_summ_mounth_end,
            'itog_summ_mounth_start': itog_summ_mounth_start,
            'money_raznica': money_raznica,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'kolvo_podkl_obj_summ': kolvo_podkl_obj_summ,
        'kolvo_otkl_obj_summ': kolvo_otkl_obj_summ,
        'rost_all': rost_all,
        'kts_itog_object_end_all': kts_itog_object_end_all,
        'kts_itog_object_all': kts_itog_object_all,
        'kts_fiz_podkl_all': kts_fiz_podkl_all,
        'kts_fiz_podkl_end_all': kts_fiz_podkl_end_all,
        'kts_count_podkl_all': kts_count_podkl_all,
        'kts_count_podkl_end_all': kts_count_podkl_end_all,
        'itog_all_end': itog_all_end,
        'itog_all_start': itog_all_start,
        'money_raznica_all': money_raznica_all,
    }

    return render(request, 'dogovornoy/kts_reports_kolvo.html', context)



@login_required
def reports_partners_techmart(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=16).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=16).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=16, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=16, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_techmart.html', context)


@login_required
def techmart_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=16, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=16, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=16, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'techmart_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[
            f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_rent_gsm']
        ws[f'M{row_num}'] = report['itog_tehnical_services']
        ws[f'N{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Tech_Mart" Fiziki {now.date()}.xlsx'

    return response


@login_required
def techmart_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=16, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=16, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=16, urik=True).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'techmart_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[
            f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['num_days']
        ws[f'J{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['reagirovanie']
        ws[f'L{row_num}'] = report['itog_rent_gsm']
        ws[f'M{row_num}'] = report['itog_tehnical_services']
        ws[f'N{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Tech_Mart" Uriki {now.date()}.xlsx'

    return response








@login_required
def reports_partners_twojoy(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=17).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=17).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=17, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=17, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))


    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_twojoy.html', context)



@login_required
def twojoy_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=17, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=17, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=17, urik=False).exclude(
        date_otkluchenia__lte=end_of_month).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, '2joy_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'M{row_num}'] = report['num_days']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "2joy" Fiziki {now.date()}.xlsx'

    return response


@login_required
def twojoy_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    print(num_days_mounth)

    partners_object_podkl = partners_object.objects.filter(company_name_id=17, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=17, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=17, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, '2joy_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['itog_fire_alarm']
        ws[f'L{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'M{row_num}'] = report['num_days']
        ws[f'N{row_num}'] = report['summ_mounth']
        ws[f'O{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'N{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "2joy" Uriki {now.date()}.xlsx'

    return response






@login_required
def reports_partners_medin(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=18).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=18).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=18, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=18, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_medin.html', context)



@login_required
def medin_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=18, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=18, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=18, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'medin_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'J{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['summ_mounth']
        ws[f'L{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'K{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Medin" Fiziki {now.date()}.xlsx'

    return response



@login_required
def medin_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=18, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=18, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=18, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'medin_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'J{row_num}'] = report['num_days']
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'K{row_num}'] = report['summ_mounth']
        ws[f'L{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'K{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Medin" Uriki {now.date()}.xlsx'

    return response





@login_required
def reports_partners_zhakitov(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days
    current_month = get_current_month_russian()
    num_days = num_days_mounth

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_of_month = datetime.strptime(end_date, '%Y-%m-%d').date()
            num_days_mounth = (end_of_month - start_of_month).days + 1
            num_days = num_days_mounth

    partners_object_podkl = partners_object.objects.filter(company_name_id=19).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    partners_kolvo_object = partners_object.objects.filter(company_name_id=19).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_ur = partners_object.objects.filter(company_name_id=19, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object_fiz = partners_object.objects.filter(company_name_id=19, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        itog_summ_mounth += summ_mounth

        reports.append({
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
        })

    context = {
        'reports': reports,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'summ_telemetria': summ_telemetria,
        'summ_rent_gsm': summ_rent_gsm,
        'summ_nabludenie': summ_nabludenie,
        'summ_reagirovanie': summ_reagirovanie,
        'summ_tehnical_services': summ_tehnical_services,
        'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
        'summ_fire_alarm': summ_fire_alarm,
        'itog_summ_mounth': itog_summ_mounth,
        'partners_kolvo_object': partners_kolvo_object,
        'partners_kolvo_object_ur': partners_kolvo_object_ur,
        'partners_kolvo_object_fiz': partners_kolvo_object_fiz,
    }

    return render(request, 'dogovornoy/reports_partners_zhakitov.html', context)



@login_required
def zhakitov_download_fiz(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=19, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=19, urik=False).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=19, urik=False).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'zakitov_download_fiz.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по физическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['num_days']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'N{row_num}'] = report['itog_rent_gsm']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Zhakitov" Fiziki {now.date()}.xlsx'

    return response


@login_required
def zhakitov_download_ur(request):
    now = timezone.now()
    start_of_month = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc).date()
    end_of_month = datetime(now.year, now.month - 1, calendar.monthrange(now.year, now.month - 1)[1],
                            tzinfo=timezone.utc).date()
    num_days_mounth = calendar.monthrange(now.year, now.month - 1)[1]  # Default to full month days

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_of_month = parse_date(start_date)
            end_of_month = parse_date(end_date)
            num_days_month = (end_of_month - start_of_month).days + 1

    partners_object_podkl = partners_object.objects.filter(company_name_id=19, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia')))
    sgs = partners_object.objects.filter(company_name_id=19, urik=True).aggregate(Count('id'))
    sgs = sgs['id__count']

    current_month = get_current_month_russian()
    current_year = get_current_year()
    partners_kolvo_object = partners_object.objects.filter(company_name_id=19, urik=True).exclude(Q(date_otkluchenia__lt=start_of_month) & Q(date_podkluchenia__lt=F('date_otkluchenia'))).aggregate(Count('id'))
    partners_kolvo_object = partners_kolvo_object.get('id__count', 0)

    reports = []
    summ_telemetria = 0
    summ_rent_gsm = 0
    summ_nabludenie = 0
    summ_reagirovanie = 0
    summ_tehnical_services = 0
    summ_sms_uvedomlenie = 0
    itog_summ_mounth = 0
    summ_fire_alarm = 0
    summ_tariff_per_mounth = 0

    for kts_instance in partners_object_podkl:
        tarif_nabludenia = None
        all_object_number = str(kts_instance.object_number) + "\\" + str(kts_instance.gsm_number)

        if kts_instance.date_otkluchenia:
            if isinstance(start_of_month, datetime):
                start_of_month = start_of_month.date()
            if isinstance(end_of_month, datetime):
                end_of_month = end_of_month.date()

            if (kts_instance.date_otkluchenia > start_of_month) and (kts_instance.date_podkluchenia < start_of_month):
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
            elif (kts_instance.date_otkluchenia >= start_of_month) and (
                    kts_instance.date_podkluchenia > kts_instance.date_otkluchenia):
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia > start_of_month:
                num_days = num_days_mounth - (kts_instance.date_podkluchenia - kts_instance.date_otkluchenia).days
            elif kts_instance.date_otkluchenia < start_of_month:
                num_days = num_days_mounth
            else:
                num_days = (kts_instance.date_otkluchenia - start_of_month).days
        else:
            if kts_instance.date_podkluchenia:
                if kts_instance.date_podkluchenia > start_of_month:
                    num_days = (end_of_month - kts_instance.date_podkluchenia).days + 1
                else:
                    num_days = num_days_mounth
            else:
                num_days = num_days_mounth

        if kts_instance.telemetria:
            itog_telemetria = int((kts_instance.company_name.telemetria / num_days_mounth) * num_days)
        else:
            itog_telemetria = 0

        if kts_instance.rent_gsm:
            if kts_instance.urik:
                itog_rent_gsm = int((kts_instance.company_name.arenda_ur / num_days_mounth) * num_days)
            else:
                itog_rent_gsm = int((kts_instance.company_name.arenda_fiz / num_days_mounth) * num_days)
        else:
            itog_rent_gsm = 0

        if kts_instance.nabludenie:
            if kts_instance.urik:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_ur / num_days_mounth) * num_days)
            else:
                itog_nabludenie = int((kts_instance.company_name.nabludenie_fiz / num_days_mounth) * num_days)
        else:
            itog_nabludenie = 0

        if kts_instance.tehnical_services:
            if kts_instance.urik:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_ur / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
            else:
                itog_tehnical_services = (kts_instance.company_name.tehnic_srv_cost_fiz / num_days_mounth) * num_days
                itog_tehnical_services = math.ceil(itog_tehnical_services) if itog_tehnical_services - math.floor(
                    itog_tehnical_services) > 0.5 else math.floor(
                    itog_tehnical_services)
        else:
            itog_tehnical_services = 0

        if kts_instance.fire_alarm:
            if kts_instance.urik:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_ur / num_days_mounth) * num_days)
            else:
                itog_fire_alarm = int((kts_instance.company_name.pozharka_fiz / num_days_mounth) * num_days)
        else:
            itog_fire_alarm = 0

        if kts_instance.urik:
            if kts_instance.tariff_per_mounth > 30:
                reagirovanie = kts_instance.tariff_per_mounth
            else:
                reagirovanie = (kts_instance.hours_mounth * kts_instance.tariff_per_mounth) / num_days_mounth * num_days
                reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                    reagirovanie)
        else:
            reagirovanie = (kts_instance.tariff_per_mounth / num_days_mounth) * num_days
            reagirovanie = math.ceil(reagirovanie) if reagirovanie - math.floor(reagirovanie) > 0.5 else math.floor(
                reagirovanie)

        if kts_instance.sms_uvedomlenie:
            if kts_instance.urik:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms_ur * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms_ur) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
            else:
                if kts_instance.sms_number:
                    itog_sms_uvedomlenie = int(
                        (kts_instance.company_name.sms * kts_instance.sms_number) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
                else:
                    itog_sms_uvedomlenie = int((kts_instance.company_name.sms) / num_days_mounth * num_days)
                    itog_sms_uvedomlenie = math.ceil(itog_sms_uvedomlenie) if itog_sms_uvedomlenie - math.floor(
                        itog_sms_uvedomlenie) > 0.5 else math.floor(
                        itog_sms_uvedomlenie)
        else:
            itog_sms_uvedomlenie = 0

        summ_mounth = itog_telemetria + itog_rent_gsm + itog_nabludenie + reagirovanie + itog_tehnical_services + itog_sms_uvedomlenie + itog_fire_alarm
        summ_telemetria += itog_telemetria
        summ_rent_gsm += itog_rent_gsm
        summ_nabludenie += itog_nabludenie
        summ_reagirovanie += reagirovanie
        summ_tehnical_services += itog_tehnical_services
        summ_sms_uvedomlenie += itog_sms_uvedomlenie
        summ_fire_alarm += itog_fire_alarm
        summ_tariff_per_mounth += kts_instance.tariff_per_mounth
        itog_summ_mounth += summ_mounth

        reports.append({
            'current_month': current_month,
            'partners_kolvo_object': partners_kolvo_object,
            'current_year': current_year,
            'kts_instance': kts_instance,
            'tarif_nabludenia': tarif_nabludenia,
            'num_days': num_days,
            'num_days_mounth': num_days_mounth,
            'itog_telemetria': itog_telemetria,
            'itog_rent_gsm': itog_rent_gsm,
            'itog_nabludenie': itog_nabludenie,
            'itog_tehnical_services': itog_tehnical_services,
            'itog_sms_uvedomlenie': itog_sms_uvedomlenie,
            'itog_fire_alarm': itog_fire_alarm,
            'reagirovanie': reagirovanie,
            'summ_mounth': summ_mounth,
            'summ_telemetria': summ_telemetria,
            'summ_rent_gsm': summ_rent_gsm,
            'summ_nabludenie': summ_nabludenie,
            'summ_reagirovanie': summ_reagirovanie,
            'summ_tehnical_services': summ_tehnical_services,
            'summ_sms_uvedomlenie': summ_sms_uvedomlenie,
            'summ_fire_alarm': summ_fire_alarm,
            'itog_summ_mounth': itog_summ_mounth,
            'summ_tariff_per_mounth': summ_tariff_per_mounth,
            'all_object_number': all_object_number,
            'sgs': sgs,
        })

    # Загрузка шаблона Excel и инициализация row_num
    template_path = os.path.join(settings.MEDIA_ROOT, 'zakitov_download_ur.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Исправляем доступ к данным
    if reports:
        first_report = reports[0]  # Берем первый отчет для заполнения заголовка
        ws[f'A{7}'] = f'АКТ сверки по юридическим лицам за {first_report["current_month"]} {first_report["current_year"]} г.'

    # Start filling in the data from row 2 (assuming row 1 is the header)
    row_num = 10
    for report in reports:
        ws[f'A{row_num}'] = report['kts_instance'].object_number
        ws[f'B{row_num}'] = report['kts_instance'].gsm_number
        ws[f'C{row_num}'] = report['kts_instance'].name_object
        ws[f'D{row_num}'] = report['kts_instance'].adres
        ws[f'E{row_num}'] = report['kts_instance'].type_object
        ws[f'F{row_num}'] = str(report['kts_instance'].vid_sign)
        ws[f'G{row_num}'] = report['kts_instance'].hours_mounth
        ws[f'H{row_num}'] = report['kts_instance'].date_podkluchenia
        ws[f'I{row_num}'] = report['kts_instance'].tariff_per_mounth
        ws[f'J{row_num}'] = report['itog_rent_gsm']
        ws[f'K{row_num}'] = report['num_days']
        ws[f'L{row_num}'] = report['reagirovanie']
        ws[f'M{row_num}'] = report['itog_sms_uvedomlenie']
        ws[f'N{row_num}'] = report['itog_rent_gsm']
        ws[f'O{row_num}'] = report['summ_mounth']
        ws[f'P{row_num}'] = report['kts_instance'].primechanie
        row_num += 1

    ws[f'C{row_num+1}'] = 'Итого'
    ws[f'O{row_num+1}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 3}'] = 'Итого охраняется:'
    ws[f'D{row_num + 3}'] = report['partners_kolvo_object']
    ws[f'C{row_num + 4}'] = f'Итого к оплате за {current_month} {current_year} г.:'
    ws[f'D{row_num + 4}'] = report['itog_summ_mounth']
    ws[f'C{row_num + 5}'] = '(В том числе НДС 12%)'
    ws[f'C{row_num + 6}'] = 'Исполнитель: бухгалтер'
    ws[f'D{row_num + 6}'] = '_________________'
    ws[f'E{row_num + 6}'] = 'Пак И.C.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TOO "Zhakitov" Urik {now.date()}.xlsx'

    return response


class CreateTaskView(CreateView):
    model = Task
    form_class = TaskForm
    clients = kts.objects.all()
    template_name = 'dogovornoy/create_task.html'
    success_url = reverse_lazy('task_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)



class TaskListView(ListView):
    model = Task
    context_object_name = 'tasks'
    template_name = 'task_list.html'
    paginate_by = 50  # Опционально, если нужна пагинация

    def get_queryset(self):
        # Используйте select_related для оптимизации загрузки связанных данных
        return Task.objects.filter(assigned_to=self.request.user).select_related('client').order_by('-created_at')



class AcceptTaskView(View):
    def post(self, request, pk):
        task = get_object_or_404(Task, pk=pk)
        task.accept_task()
        return redirect(reverse('task_list'))


class CompleteTaskView(View):
    def post(self, request, pk):
        note = request.POST.get('note', '')
        task = get_object_or_404(Task, pk=pk)
        task.complete_task(note)
        return redirect(reverse('task_list'))

# Телеграм бот для техников

pending_results = {}

class CreateTechnicalTaskView(View):
    def post(self, request, *args, **kwargs):
        technician_id = request.POST.get('technician_id')
        client_object_id = request.POST.get('client_object_id')
        ekcbase_object_id = request.POST.get('client_object_id')
        note = request.POST.get('note')
        reason = request.POST.get('reason')  # Получаем строку с причинами через запятую

        # Получаем технику
        technician = get_object_or_404(User, pk=technician_id)

        # Создаем задачу
        task = TechnicalTask.objects.create(
            technician=technician,
            sender=request.user,
            client_object_id=client_object_id,
            ekcbase_object_id=ekcbase_object_id,
            reason=reason,  # Сохраняем строку с причинами
            note=note
        )

        # Отправляем уведомление в Telegram (если нужно)
        send_telegram_message(technician, task)

        return redirect('technical_task_list')


def get_card_from_third_db(card_id):
    try:
        card = Cards.objects.using('third_db').get(cardid=card_id)
        return card
    except Cards.DoesNotExist:
        return None


def get_zones_from_third_db(card_id):
    try:
        zones = Zones.objects.using('third_db').filter(cardid=card_id).select_related('sectionid')
        return zones
    except Zones.DoesNotExist:
        return None


def get_card_from_asuekc(card_id):
    try:
        card_asuekc = GuardedObjects.objects.using('asu_ekc').get(pk=card_id)
        return card_asuekc
    except Cards.DoesNotExist:
        return None


bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)

# Глобальная переменная для хранения последнего SN (можно заменить на БД или кэш)
last_event_sn = {}


def message_handler(update, context):
    user_id = update.message.from_user.id
    text = update.message.text

    # Проверяем, ожидает ли пользователь ввода результата
    if user_id in pending_results:
        task_id = pending_results[user_id]
        try:
            # Проверяем, существует ли задача
            task = TechnicalTask.objects.get(pk=task_id)

            # Сохраняем результат
            task.result = text  # Поле для хранения результата (добавьте его в модель, если еще нет)
            task.save()

            # Уведомляем пользователя об успешном сохранении
            update.message.reply_text(f"Результат для заявки #{task_id} успешно сохранен.")

            # Удаляем из ожидающих
            del pending_results[user_id]
        except TechnicalTask.DoesNotExist:
            # Уведомляем, что задача не найдена
            update.message.reply_text(f"Заявка #{task_id} больше не существует в системе.")

            # Удаляем из списка ожидающих
            del pending_results[user_id]
        except Exception as e:
            # Обрабатываем другие ошибки
            update.message.reply_text(f"Ошибка при сохранении результата: {e}")
    else:
        # Если пользователь не в ожидании ввода
        update.message.reply_text("Я не ожидал от вас результата. Попробуйте снова.")


def button_handler(update, context):
    global last_event_sn
    query = update.callback_query
    data = query.data.split('_')

    if len(data) < 2:
        query.edit_message_text(text="Ошибка: Некорректные данные.")
        return

    action = data[0]
    task_id = data[-1]

    try:
        task = TechnicalTask.objects.get(pk=task_id)
        if action == "select":
            # Получаем зоны клиента
            zones = get_zones_from_third_db(task.client_object_id)
            if zones.exists():
                message = "Зоны клиента:\n"
                for zone in zones:
                    message += f"Получено: {timezone.now().strftime('%H:%M:%S')}\n\n"
                    message += f"Раздел: {zone.sectionid.sectionname}, Зона: {zone.zonenumber} - {zone.info}\n"
            else:
                message = "Зоны не найдены."

            keyboard = [
                [InlineKeyboardButton(text="🔴 Показать заявку", callback_data=f"task_show_{task_id}")],
                [InlineKeyboardButton(text="🟢 Вывести события", callback_data=f"task_module_{task_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            query.edit_message_text(text=message, reply_markup=reply_markup)

        elif action == "task" and data[1] == "show":
            send_telegram_message(task.technician, task)
            query.edit_message_text(text="Заявка была повторно отправлена.")

        elif action == "task" and data[1] == "module":
            # Отображение событий
            card = get_card_from_third_db(task.client_object_id)
            module_number = card.unitnumber
            alarms = execute_stored_procedure(module_number)

            if alarms:
                print(alarms)
                message = f"Результаты для модуля {module_number}:\n"
                message += f"Получено: {timezone.now().strftime('%H:%M:%S')}\n\n"
                for row in reversed(alarms):
                    razdel = row[3] if row[3] else '-'
                    zona_user = row[4] if row[4] else '-'
                    event_str = row[6]
                    date_event = row[8].strftime('%d-%m-%Y %H:%M:%S') if row[8] else 'Нет данных'
                    gprs_quality = row[15]
                    message += f"Раздел: {razdel}, Зона/Польз: {zona_user}, Событие: {event_str}, Дата: {date_event}, Качество: {gprs_quality}\n\n"
                    last_event_sn[task_id] = alarms[0][1]
                # Сохраняем последний SN из полученных данных
                print(f"SN из последнего события + {last_event_sn[task_id]}")
            else:
                message = "События не найдены."

            keyboard = [
                [InlineKeyboardButton(text="🔴 Показать заявку", callback_data=f"task_show_{task_id}")],
                [InlineKeyboardButton(text="🟢 Вывести зоны клиента", callback_data=f"select_task_{task_id}")],
                [InlineKeyboardButton(text="🟢 Обновить события", callback_data=f"update_task_{task_id}")]
            ]

            reply_markup = InlineKeyboardMarkup(keyboard)
            query.edit_message_text(text=message, reply_markup=reply_markup)

        elif action == "update":
            # Проверка на наличие сохраненного `SN` для обновления
            card = get_card_from_third_db(task.client_object_id)
            module_number = card.unitnumber
            new_alarms = execute_stored_procedure(module_number)

            if new_alarms:
                message = f"Обновленные события для модуля {module_number}:\n"
                message += f"Обновлено: {timezone.now().strftime('%H:%M:%S')}\n\n"
                for row in reversed(new_alarms):
                    razdel = row[3] if row[3] else '-'
                    zona_user = row[4] if row[3] else '-'
                    event_str = row[6]
                    date_event = row[8].strftime('%d-%m-%Y %H:%M:%S') if row[8] else 'Нет данных'
                    gprs_quality = row[15]
                    message += f"Раздел: {razdel}, Зона/Польз: {zona_user}, Событие: {event_str}, Дата: {date_event}, Качество: {gprs_quality}\n\n"
                    last_event_sn[task_id] = new_alarms[0][1]

                # Обновляем последний SN для последующих обновлений
                # print(f"Обновленный last_event_sn = {last_event_sn[task_id]}")
            else:
                message = "Нет новых событий для модуля."
                message += f"Обновлено: {timezone.now().strftime('%H:%M:%S')}\n\n"

            # Добавляем кнопки
            keyboard = [
                [InlineKeyboardButton(text="🔴 Показать заявку", callback_data=f"task_show_{task_id}")],
                [InlineKeyboardButton(text="🟢 Вывести зоны клиента", callback_data=f"select_task_{task_id}")],
                [InlineKeyboardButton(text="🟢 Обновить события", callback_data=f"update_task_{task_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

            # Проверка перед редактированием сообщения
            if query.message.text != message:
                query.edit_message_text(text=message, reply_markup=reply_markup)
            else:
                print("Содержимое сообщения и разметка не изменились, обновление пропущено.")

        elif action == "arrival":
            card = get_card_from_third_db(task.client_object_id)
            if card:
                # Сохраняем текущее значение workstation перед изменением
                task.previous_workstation = card.workstation
                task.arrival_time = timezone.now()
                task.save()

                # Изменяем workstation на 3 (техническое обслуживание)
                card.workstation = 3
                card.save(using='third_db')  # Сохраняем изменение в базе данных `third_db`

                message = f"Время прибытия на объект установлено: {task.arrival_time.strftime('%d-%m-%Y %H:%M:%S')}\n" \
                          f"Клиент временно переведен на техническое обслуживание."
                keyboard = [
                    [InlineKeyboardButton(text="🔴 Показать заявку", callback_data=f"task_show_{task_id}")],
                    [InlineKeyboardButton(text="🟢 Вывести зоны клиента", callback_data=f"select_task_{task_id}")],
                    [InlineKeyboardButton(text="🟢 Вывести события", callback_data=f"task_module_{task_id}")],
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                query.edit_message_text(text=message, reply_markup=reply_markup)
            else:
                query.edit_message_text(text="Клиент не найден.")
        elif action == "completion":
            card = get_card_from_third_db(task.client_object_id)
            if card and task.previous_workstation is not None:
                # Возвращаем значение workstation к предыдущему состоянию
                card.workstation = task.previous_workstation
                card.save(using='third_db')

                # Сбрасываем сохраненное значение, так как задача завершена
                task.previous_workstation = None
                task.completion_time = timezone.now()
                task.save()

                message = f"Время завершения заявки установлено: {task.completion_time.strftime('%d-%m-%Y %H:%M:%S')}\n" \
                          f"Клиент переведен в прежний режим работы."
                keyboard = [
                    [InlineKeyboardButton(text="🔴 Заполнить результат", callback_data=f"result_{task_id}")],
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                query.edit_message_text(text=message, reply_markup=reply_markup)
            else:
                query.edit_message_text(text="Ошибка: невозможно восстановить состояние клиента.")
        elif action == "result":
            # Сохраняет результат выполнения от техника
            pending_results[query.from_user.id] = task_id

            message = f"Пожалуйста, отправьте текстовое сообщение с результатом для заявки #{task_id}."
            query.edit_message_text(text=message)

    except Exception as e:
        print(f"Ошибка при обработке запроса: {e}")
        query.edit_message_text(text=f"Произошла ошибка при получении данных: {e}")
        send_telegram_message(task.technician, task)
        query.edit_message_text(text="Заявка была повторно отправлена.")


# Функция для отправки сообщения с кнопками
def send_telegram_message(technician, task):
    card = get_card_from_third_db(task.client_object_id)
    bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
    message = f"Номер объекта: {card.otisnumber}\n"
    message += f"Новая заявка для {technician.first_name} {technician.last_name} ({technician.username}):\n"
    message += f"Наименование клиента: {card.objectname}\n"
    message += f"Номер модуля: {card.unitnumber}\n"
    message += f"Адрес: {card.info}\n"
    message += f"Телефон: {card.phones}\n"
    message += f"Причина: {task.reason}\n"
    message += f"Примечание к заявке: {task.note}\n"

    # Формируем inline-кнопку для заявки
    if task.arrival_time:
        keyboard = [
            [InlineKeyboardButton(text="🔴 Завершить заявку", callback_data=f"completion_task_{task.id}")],
        ]
    else:
        keyboard = [
            [InlineKeyboardButton(text="🟡 Прибыл на объект", callback_data=f"arrival_task_{task.id}")],
        ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Отправляем сообщение с кнопкой
    try:
        bot.send_message(
            chat_id=technician.userprofile.telegram_id,
            text=message,
            reply_markup=reply_markup
        )
        print("Сообщение и кнопка успешно отправлены")
    except Exception as e:
        print(f"Ошибка при отправке сообщения в Telegram: {e}")


@csrf_exempt
def telegram_webhook(request):
    try:
        data = json.loads(request.body.decode('utf-8'))

        update = Update.de_json(data, bot)
        dispatcher = Dispatcher(bot, None, workers=0)
        dispatcher.add_handler(CallbackQueryHandler(button_handler))
        dispatcher.add_handler(MessageHandler(Filters.text & ~Filters.command, message_handler))

        dispatcher.process_update(update)
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        print(f"Ошибка в webhook: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def module_button_handler(update, context):
    query = update.callback_query
    task_id = query.data.split('_')[-1]  # Получаем task_id из callback_data

    try:
        # Получаем объект заявки из базы данных
        task = TechnicalTask.objects.get(pk=task_id)

        # Получаем информацию о клиенте (модуле) через client_object_id
        card = get_card_from_third_db(task.client_object_id)

        # Если объект найден, извлекаем номер модуля (unitnumber)
        if card and card.unitnumber:
            module_number = card.unitnumber
            print(f"Отладка module_number в module_button_handler: {module_number}")
            result = execute_stored_procedure(module_number)

            # Формируем сообщение для отправки в Telegram
            if result:
                message = f"Результаты для модуля {module_number}:\n"
                for row in result:
                    message += f"{row}\n"
            else:
                message = f"Ошибка при выполнении запроса для модуля {module_number}."
        else:
            message = "Модуль не найден для данного клиента."

        # Отправляем сообщение с результатами выполнения процедуры
        query.edit_message_text(text=message)

    except Exception as e:
        print(f"Ошибка при обработке запроса: {e}")
        query.edit_message_text(text="Произошла ошибка при получении данных.")


def execute_stored_procedure(module_number):
    try:
        # Открываем соединение с базой данных 'third_db'
        with connections['third_db'].cursor() as cursor:
            # Определяем переменные
            date_now = timezone.now() - timezone.timedelta(hours=0.5)  # Дата от текущего времени минус 1 час
            date_minus_3_days = date_now

            # Отладочная информация
            print(f"Выполнение хранимой процедуры с параметрами: дата - {date_minus_3_days}, модуль - {module_number}")

            # Выполняем хранимую процедуру
            cursor.execute("""
                DECLARE @D datetime;
                DECLARE @I int;
                DECLARE @M int;
                SET @D = %s;
                SET @M = %s;
                SET @I = 0;
                EXECUTE [sp_GSM2MSG_MODUL_ADMIN] @D, @I, @M;
            """, [date_minus_3_days, module_number])

            # Получаем результаты
            result = cursor.fetchall()
            print(f"Дата {date_minus_3_days}:")
            return result
    except Exception as e:
        print(f"Ошибка при выполнении хранимой процедуры: {e}")
        return None


# Функция для выполнения хранимой процедуры с последним ID события
def execute_stored_procedure_with_last_id(module_number, last_event_id):
    try:
        with connections['third_db'].cursor() as cursor:
            # Задаем параметры для выполнения процедуры
            date_now = 0  # Чтобы запрос возвращал только новые данные
            cursor.execute("""
                DECLARE @D datetime;
                DECLARE @I int;
                DECLARE @M int;
                SET @M = %s;
                SET @D = %s;
                SET @I = %s;
                EXECUTE [sp_GSM2MSG_MODUL_ADMIN] @D, @I, @M;
            """, [module_number, date_now, last_event_id])

            # Получаем результаты
            result = cursor.fetchall()
            print(f"Новые события для модуля {module_number}: {result}")
            return result
    except Exception as e:
        print(f"Ошибка при выполнении процедуры обновления событий: {e}")
        return None



def TechniciansAPIView(request):
    technicians = User.objects.filter(userprofile__department__icontains="Техник")  # Фильтрация списка техников
    data = [
        {'id': technician.id,
         "first_name": technician.first_name or "",
         "last_name": technician.last_name or "",
         "username": technician.username,
         }
         for technician in technicians
    ]
    return JsonResponse(data, safe=False)


def TaskReasonsAPIView(request):
    reasons = TaskReason.objects.all()
    data = [{'id': reason.id, 'reason': reason.reason} for reason in reasons]
    return JsonResponse(data, safe=False)


class TechnicalTaskListView(ListView):
    model = TechnicalTask
    template_name = 'dogovornoy/technical_task_list.html'
    context_object_name = 'tasks'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().filter(completion_time__isnull=True)
        form = self.get_filter_form()

        # Фильтр по ID клиента
        client_object_id = form.cleaned_data.get('client_object_id')
        if client_object_id:
            queryset = queryset.filter(client_object_id=client_object_id)

        # Фильтр по технику
        technician = form.cleaned_data.get('technician')
        if technician:
            queryset = queryset.filter(technician=technician)

        # Фильтр по диапазону дат
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(sent_time__range=[start_date, end_date])

        queryset = queryset.annotate(
             sort_priority = Case(
                 When(arrival_time__isnull=False, then=Value(0)),
                 default=Value(1),
             )
        ).order_by('sort_priority', '-arrival_time', '-sent_time')

            # Дополняем задачи информацией из `Cards`
        tasks_with_card_info = []
        for task in queryset:
            card = get_card_from_third_db(task.client_object_id)
            task.card_info = card  # Добавляем объект `Cards` к задаче
            tasks_with_card_info.append(task)

        return tasks_with_card_info

    def get_filter_form(self):
        form = TechnicalTaskFilterForm(self.request.GET)
        if form.is_valid():
            return form
        return TechnicalTaskFilterForm()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        queryset = self.get_queryset()
        paginator = Paginator(queryset, self.paginate_by)

        # Получение номера страницы
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Копируем параметры запроса для использования в пагинации
        params = self.request.GET.copy()
        if 'page' in params:
            del params['page']
        pagination_url = self.request.path + '?' + urlencode(params)

        # Добавляем объекты пагинации в контекст
        context['tasks'] = page_obj
        context['pagination_url'] = pagination_url
        context['filter_form'] = self.get_filter_form()
        return context



class ArchiveTechnicalTaskListView(ListView):
    model = TechnicalTask
    template_name = 'dogovornoy/archive_technical_task_list.html'
    context_object_name = 'tasks'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().filter(completion_time__isnull=False)
        form = self.get_filter_form()

        # Фильтр по ID клиента
        client_object_id = form.cleaned_data.get('client_object_id')
        if client_object_id:
            queryset = queryset.filter(client_object_id=client_object_id)

        # Фильтр по технику
        technician = form.cleaned_data.get('technician')
        if technician:
            queryset = queryset.filter(technician=technician)

        # Фильтр по диапазону дат
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(sent_time__range=[start_date, end_date])

        queryset = queryset.annotate(
             sort_priority = Case(
                 When(arrival_time__isnull=False, then=Value(0)),
                 default=Value(1),
             )
        ).order_by('sort_priority', '-arrival_time', '-sent_time')

            # Дополняем задачи информацией из `Cards`
        tasks_with_card_info = []
        for task in queryset:
            card = get_card_from_third_db(task.client_object_id)
            task.card_info = card  # Добавляем объект `Cards` к задаче
            tasks_with_card_info.append(task)

        return tasks_with_card_info

    def get_filter_form(self):
        form = TechnicalTaskFilterForm(self.request.GET)
        if form.is_valid():
            return form
        return TechnicalTaskFilterForm()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        queryset = self.get_queryset()
        paginator = Paginator(queryset, self.paginate_by)

        # Получение номера страницы
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Копируем параметры запроса для использования в пагинации
        params = self.request.GET.copy()
        if 'page' in params:
            del params['page']
        pagination_url = self.request.path + '?' + urlencode(params)

        # Добавляем объекты пагинации в контекст
        context['tasks'] = page_obj
        context['pagination_url'] = pagination_url
        context['filter_form'] = self.get_filter_form()
        return context



class DisconnectedObjectsView(ListView):
    template_name = 'dogovornoy/disconnected_objects.html'
    context_object_name = 'disconnected_objects'
    paginate_by = 100  # Настройка пагинации

    def get_queryset(self):
        now = timezone.now()
        first_day_of_prev_month = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_day_of_prev_month = now.replace(day=1) - timedelta(days=1)

        # Отключенные объекты ваших клиентов
        client_disconnected = kts.objects.filter(
            date_otklulchenia__gte=first_day_of_prev_month,
            date_otklulchenia__lte=last_day_of_prev_month
        ).annotate(client_type=ExpressionWrapper(Value('Наш клиент'), output_field=CharField()))

        # Отключенные объекты клиентов партнеров
        partner_disconnected = partners_object.objects.filter(
            date_otkluchenia__gte=first_day_of_prev_month,
            date_otkluchenia__lte=last_day_of_prev_month
        ).annotate(client_type=ExpressionWrapper(Value('Клиент партнера'), output_field=CharField()))

        # Объединяем результаты в один список
        disconnected_objects = list(client_disconnected) + list(partner_disconnected)

        return disconnected_objects

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paginator = self.get_paginator(self.get_queryset(), self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['disconnected_objects'] = page_obj
        return context



def export_disconnected_objects(request):
    # Получаем текущий месяц
    now_time = timezone.now()
    first_day_of_prev_month = (now_time.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_day_of_prev_month = now_time.replace(day=1) - timedelta(days=1)

    # Отключенные объекты ваших клиентов
    client_disconnected = kts.objects.filter(
        date_otklulchenia__gte=first_day_of_prev_month,
        date_otklulchenia__lte=last_day_of_prev_month
    ).annotate(client_type=Value('Наш клиент', output_field=CharField()))

    # Отключенные объекты клиентов партнеров
    partner_disconnected = partners_object.objects.filter(
        date_otkluchenia__gte=first_day_of_prev_month,
        date_otkluchenia__lte=last_day_of_prev_month
    ).annotate(client_type=Value('Клиент партнера', output_field=CharField()))

    # Объединяем результаты
    disconnected_objects = list(client_disconnected)

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отключенные объекты'

    # Заголовки для столбцов
    headers = ['Компания', 'Номер договора', 'Номер объекта', 'Наименование клиента', 'Адрес', 'Дата подключения', 'Дата отключения', 'Примечание','База']
    ws.append(headers)


    # Заполнение данными
    for obj in disconnected_objects:
        # Проверяем, является ли поле объектом модели и преобразуем его в строку
        company_name = str(obj.company_name) if obj.company_name else ''
        dogovor_number = str(getattr(obj, 'dogovor_number', ''))
        object_number = str(obj.object_number) if obj.object_number else ''
        klient_name = str(getattr(obj, 'klient_name', obj.name_object)) if obj.name_object else ''
        adres = str(obj.adres) if obj.adres else ''
        date_podkluchenia = str(obj.date_podkluchenia) if obj.date_podkluchenia else ''
        date_otkluchenia_str = str(obj.date_otklulchenia)
        primechanie = str(obj.primechanie) if obj.primechanie else '-'
        client_type = str(obj.client_type) if obj.client_type else ''

        row = [company_name, dogovor_number, object_number, klient_name, adres, date_podkluchenia, date_otkluchenia_str,
               primechanie, client_type]
        ws.append(row)

    # Генерация ответа для скачивания файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=disconnected_objects.xlsx'
    wb.save(response)

    return response



def export_disconnected_objects_partners(request):
    # Получаем текущий месяц
    now_time = timezone.now()
    first_day_of_prev_month = (now_time.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_day_of_prev_month = now_time.replace(day=1) - timedelta(days=1)

    # Отключенные объекты ваших клиентов
    client_disconnected = kts.objects.filter(
        date_otklulchenia__gte=first_day_of_prev_month,
        date_otklulchenia__lte=last_day_of_prev_month
    ).annotate(client_type=Value('Наш клиент', output_field=CharField()))

    # Отключенные объекты клиентов партнеров
    partner_disconnected = partners_object.objects.filter(
        date_otkluchenia__gte=first_day_of_prev_month,
        date_otkluchenia__lte=last_day_of_prev_month
    ).annotate(client_type=Value('Клиент партнера', output_field=CharField()))

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отключенные объекты'

    # Заголовки для столбцов
    headers = ['Компания', 'Номер объекта', 'Наименование клиента', 'Адрес', 'Дата подключения', 'Дата отключения', 'Примечание','База']
    ws.append(headers)


    # Заполнение данными
    for obj in partner_disconnected:
        company_name = str(obj.company_name) if obj.company_name else ''
        object_number = str(obj.object_number) if obj.object_number else ''
        klient_name = str(getattr(obj, 'klient_name', obj.name_object)) if obj.name_object else ''
        adres = str(obj.adres) if obj.adres else ''
        date_podkluchenia = str(obj.date_podkluchenia) if obj.date_podkluchenia else ''
        date_otkluchenia_str = str(obj.date_otkluchenia)
        primechanie = str(obj.primechanie) if obj.primechanie else '-'
        client_type = str(obj.client_type) if obj.client_type else ''

        row = [company_name, object_number, klient_name, adres, date_podkluchenia, date_otkluchenia_str,
               primechanie, client_type]
        ws.append(row)

    # Генерация ответа для скачивания файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=disconnected_objects_partners.xlsx'
    wb.save(response)

    return response



class ArchiveTaskListView(ListView):
    model = Task
    template_name = 'dogovornoy/archive_task_list.html'  # Указываем шаблон для архива заявок
    context_object_name = 'tasks'
    paginate_by = 10  # Пагинация, если требуется

    def get_queryset(self):
        user_id = self.request.user.id
        # Фильтруем завершенные заявки, связанные с пользователем
        return Task.objects.filter(
            Q(assigned_to_id=user_id) | Q(created_by_id=user_id),  # Фильтруем по id пользователя
            completed_at__isnull=False  # Только завершенные заявки
        ).order_by('-completed_at')



def add_skaldgsm(request):
    if request.method == 'POST':
        form = SkaldGSMForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse('skladgsm_list'))  # Замените на нужный URL после сохранения
    else:
        form = SkaldGSMForm()
    return render(request, 'dogovornoy/add_skaldgsm.html', {'form': form})



def skladgsm_list(request):
    skladgsm_items = SkaldGSM2.objects.all()
    return_form = DateBackGSMForm()  # Форма для модального окна
    return render(request, 'dogovornoy/skladgsm_list.html', {'skladgsm_items': skladgsm_items, 'return_form': return_form})


def skladgsm_return(request, pk):
    item = get_object_or_404(SkaldGSM2, pk=pk)
    if request.method == 'POST':
        form = DateBackGSMForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('skladgsm_list')  # Redirect back to the list view
    else:
        form = DateBackGSMForm(instance=item)
    return render(request, 'skladgsm_return.html', {'form': form, 'item': item})



def export_kts_to_exel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "База договоров"

    headers = [
        "Номер УДВ", "Дата и Кем выдано", "Компания", "№ дог.", "Дата заключения",
        "Наличие Договора", "Мат.отв", "Акты ТУ", "Время реагирования",
        "Реагирование не более", "Условия договора", "Наименование Клиента",
        "Наименование объекта", "Адрес объекта", "ИИН/БИН", "Телефон",
        "Вид сигнализации", "Юридическое лицо", "Часы по договору", "Алсеко",
        "Абон.плата", "№ объекта", "№ передатчика/GSM", "Стоимость РПО",
        "Дата подключения", "Дата отключения", "Дата изменения",
        "Группа реагирования", "Электронный адрес", "Вид РПО", "Примечание",
        "Агентские", "ИИК", "БИК", "БАНК", "Режим работы",
        "Имя директора сокращенное", "Имя директора полное", "Должность директора",
        "Учередительные документы", "Юридический адрес"
    ]

    sheet.append(headers)

    for obj in kts.objects.all():
        row = [
            obj.udv_number, obj.date_udv, obj.company_name.polnoe_name,
            obj.dogovor_number, obj.data_zakluchenia, obj.nalichiye_dogovora,
            obj.mat_otv, obj.act_ty, obj.time_reag, obj.time_reag_nebol,
            obj.yslovie_dogovora, obj.klient_name, obj.name_object,
            obj.adres, obj.iin_bin, obj.telephone,
            obj.vid_sign.name_sign, obj.urik, obj.chasi_po_dog,
            obj.dop_uslugi, obj.abon_plata, obj.object_number, obj.peredatchik_number,
            obj.stoimost_rpo, obj.date_podkluchenia, obj.date_otklulchenia,
            obj.date_izmenenia, obj.gruppa_reagirovania, obj.email, obj.vid_rpo,
            obj.primechanie, obj.agentskie, obj.iik, obj.bik, obj.bank,
            obj.rezhim_raboti, obj.fio_direktor_sokr, obj.fio_direktor_polnoe,
            obj.dolznost, obj.ucereditel_doc, obj.urik_adress
        ]
        sheet.append(row)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="baza_dogovorov.xlsx"'
    workbook.save(response)

    return response



def export_partners_to_excel(request):
    # Создаем новый Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Объекты партнеров"

    # Заголовки для столбцов
    headers = [
        "Номер объекта", "Номер GSM", "Наименование клиента", "Адрес", "Тип объекта",
        "Вид сигнализации", "Часы в месяц", "Дата подключения",
        "Тариф за мониторинг и реагирование в месяц", "Тех.обслуживание", "Аренда GSM",
        "Пожарная сигнализация", "Телеметрия", "Наблюдение", "SMS уведомление",
        "SMS кол-во номеров", "Кол-во дней", "Примечание", "Экипаж",
        "Юридическое лицо", "Партнеры", "Дата отключения", "Прочее"
    ]

    # Добавляем заголовки в первую строку
    sheet.append(headers)

    # Получаем данные из модели partners_object
    for obj in partners_object.objects.all():
        row = [
            obj.object_number, obj.gsm_number, obj.name_object, obj.adres, obj.type_object,
            obj.vid_sign.name_sign, obj.hours_mounth, obj.date_podkluchenia,
            obj.tariff_per_mounth, obj.tehnical_services, obj.rent_gsm,
            obj.fire_alarm, obj.telemetria, obj.nabludenie, obj.sms_uvedomlenie,
            obj.sms_number, obj.kolvo_day, obj.primechanie,
            obj.ekipazh.ekipazh_name if obj.ekipazh else None, obj.urik,
            obj.company_name.polnoe_name, obj.date_otkluchenia,
            obj.prochee
        ]
        sheet.append(row)

    # Настраиваем ответ для скачивания файла
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="partners_objects.xlsx"'
    workbook.save(response)

    return response




